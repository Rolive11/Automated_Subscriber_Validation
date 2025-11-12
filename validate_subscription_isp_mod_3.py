import csv
import psycopg2
from psycopg2.extras import RealDictCursor
import sys
import os
import shutil
from datetime import datetime
import smtplib
import ssl
import email
import glob
import subprocess
import math
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests
import googlemaps
from time import time, sleep
import re
import json

# Load environment variables from .env file if it exists
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    # dotenv not installed, rely on system environment variables
    pass

# Email configuration constants
EMERGENCY_EMAIL = 'rolive@regulatorysolutions.us'
EMAIL_CONFIG_PATH = '/var/www/broadband/src/config/email_config.json'

# Global email config cache
_email_config_cache = None
_email_config_error = None


def load_email_config():
    """Load email configuration from JSON file with caching and fallback to defaults."""
    global _email_config_cache, _email_config_error

    # Return cached config if available
    if _email_config_cache is not None:
        return _email_config_cache, None

    # Return cached error if we already tried and failed
    if _email_config_error is not None:
        return get_default_email_config(), _email_config_error

    try:
        with open(EMAIL_CONFIG_PATH, 'r') as f:
            config = json.load(f)

        # Validate required fields
        required_fields = ['from_address', 'admin_email', 'bcc_addresses', 'smtp_user']
        missing_fields = [field for field in required_fields if field not in config]

        if missing_fields:
            error_msg = f"Missing required fields in email config: {', '.join(missing_fields)}"
            _email_config_error = error_msg
            with open('validate_subs.log', 'a') as f:
                print(f'[EMAIL CONFIG ERROR] {error_msg}\n', file=f)
            return get_default_email_config(), error_msg

        # Parse BCC addresses and ensure emergency email is included
        bcc_list = [addr.strip() for addr in config['bcc_addresses'].split(',')]
        if EMERGENCY_EMAIL not in bcc_list:
            bcc_list.append(EMERGENCY_EMAIL)
        config['bcc_list'] = bcc_list

        _email_config_cache = config
        with open('validate_subs.log', 'a') as f:
            print(f'[EMAIL CONFIG] Successfully loaded from {EMAIL_CONFIG_PATH}\n', file=f)

        return config, None

    except FileNotFoundError:
        error_msg = f"Email config file not found: {EMAIL_CONFIG_PATH}"
        _email_config_error = error_msg
        with open('validate_subs.log', 'a') as f:
            print(f'[EMAIL CONFIG ERROR] {error_msg}\n', file=f)
        return get_default_email_config(), error_msg

    except json.JSONDecodeError as e:
        error_msg = f"Invalid JSON in email config file: {str(e)}"
        _email_config_error = error_msg
        with open('validate_subs.log', 'a') as f:
            print(f'[EMAIL CONFIG ERROR] {error_msg}\n', file=f)
        return get_default_email_config(), error_msg

    except Exception as e:
        error_msg = f"Unexpected error loading email config: {str(e)}"
        _email_config_error = error_msg
        with open('validate_subs.log', 'a') as f:
            print(f'[EMAIL CONFIG ERROR] {error_msg}\n', file=f)
        return get_default_email_config(), error_msg


def get_default_email_config():
    """Return hard-coded default email configuration."""
    return {
        'from_address': 'info@regulatorysolutions.us',
        'admin_email': EMERGENCY_EMAIL,
        'bcc_addresses': EMERGENCY_EMAIL,
        'bcc_list': [EMERGENCY_EMAIL],
        'smtp_user': 'info@regulatorysolutions.us'
    }


def send_emergency_notification(error_message, intended_recipient, context_info):
    """Send emergency notification about email config failure to hard-coded emergency email."""
    try:
        port = 465
        context = ssl.create_default_context()

        message = MIMEMultipart()
        message["From"] = 'info@regulatorysolutions.us'
        message["To"] = EMERGENCY_EMAIL
        message["Subject"] = 'WARNING: Email Config Error - validate_subscription_isp'

        body = f"""WARNING: Email Configuration Error

The email configuration file could not be loaded:
File: {EMAIL_CONFIG_PATH}
Error: {error_message}

Fallback action taken: Email WAS SENT to {intended_recipient} using hard-coded defaults.

Context:
{context_info}

Note: Only {EMERGENCY_EMAIL} was copied on the email (BCC).

Please fix the email_config.json file to restore full functionality.
"""

        message.attach(MIMEText(body, "plain"))
        text = message.as_string()

        smtp_password = os.getenv('SMTP_PASSWORD')
        if not smtp_password:
            with open('validate_subs.log', 'a') as f:
                print(f'[EMERGENCY EMAIL ERROR] Cannot send emergency notification - SMTP_PASSWORD not set\n', file=f)
            return

        with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
            server.login('info@regulatorysolutions.us', smtp_password)
            server.sendmail('info@regulatorysolutions.us', EMERGENCY_EMAIL, text)

        with open('validate_subs.log', 'a') as f:
            print(f'[EMERGENCY EMAIL] Sent emergency notification to {EMERGENCY_EMAIL}\n', file=f)

    except Exception as e:
        with open('validate_subs.log', 'a') as f:
            print(f'[EMERGENCY EMAIL ERROR] Failed to send emergency notification: {str(e)}\n', file=f)


def truncate(f, n):
    return math.floor(f * 10 ** n) / 10 ** n


def geoCode(address_or_zipcode):

    lat, lng = None, None
    api_key = os.getenv('GOOGLE_MAPS_API_KEY')
    if not api_key:
        raise ValueError("GOOGLE_MAPS_API_KEY environment variable not set")
    gmaps = googlemaps.Client(key=api_key)

    geocode_result = gmaps.geocode(address_or_zipcode)
    # print(geocode_result)
    results = geocode_result[0]['geometry']['location']
    lat = results['lat']
    lng = results['lng']
    # print('lat lng ' + str(lat) + ' ' + str(lng))

    # base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    # endpoint = f"{base_url}?address={address_or_zipcode}&amp;key={api_key}"
    # see how our endpoint includes our API key? Yes this is yet another reason to restrict the key
    # r = requests.get(endpoint)
    # print(r)
    # print("status code " + str(r.status_code))
    # if r.status_code not in range(200, 299):
    #    print('status code not in range')
    #    return None, None
    # try:
    #    '''
    #    This try block incase any of our inputs are invalid. This is done instead
    #    of actually writing out handlers for all kinds of responses.
    #    '''
    # print('got geocode response ')
    # results = r.json()['results'][0]
    # print("results " + results)
    # lat = results['geometry']['location']['lat']
    # lng = results['geometry']['location']['lng']
    # except:
    #    print('pass')
    #    pass  '''
    return lat, lng


def sendEmail(customer, name, emessage, attachment_path=None, subject=None):
    """Send email to customer with optional file attachment(s).

    Args:
        attachment_path: Can be a single path string or a list of paths
    """
    with open('validate_subs.log', 'a') as f:
        print(f'[SEND EMAIL] Preparing to send email to customer: {customer}\n', file=f)

    # Load email configuration
    email_config, config_error = load_email_config()

    # If config failed, send emergency notification
    if config_error:
        send_emergency_notification(
            config_error,
            customer,
            f"Email type: Customer notification\nRecipient: {customer}\nSubject: {subject or 'Subscriber File Processing Update'}"
        )

    port = 465  # For SSL
    # Create a secure SSL context
    context = ssl.create_default_context()

    message = MIMEMultipart()
    message["From"] = email_config['from_address']
    message["To"] = customer
    default_subject = 'Automated Message - Subscriber File Processing Update'
    message["Subject"] = subject if subject else default_subject
    # Add BCC recipients from config
    message["Bcc"] = ', '.join(email_config['bcc_list'])

    with open('validate_subs.log', 'a') as f:
        print(f'[SEND EMAIL] BCC list: {", ".join(email_config["bcc_list"])}\n', file=f)

    # Add body to email
    message.attach(MIMEText(emessage, "plain"))

    # Handle both single attachment path (string) and multiple paths (list)
    attachment_paths = []
    if attachment_path:
        if isinstance(attachment_path, list):
            attachment_paths = attachment_path
        else:
            attachment_paths = [attachment_path]

    # Add file attachments if provided
    for att_path in attachment_paths:
        if att_path and os.path.exists(att_path):
            try:
                with open(att_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                # Encode file in ASCII characters to send by email
                encoders.encode_base64(part)

                # Add header as key/value pair to attachment part
                filename = os.path.basename(att_path)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {filename}",
                )

                # Add attachment to message
                message.attach(part)

                with open('validate_subs.log', 'a') as f:
                    print(f'Added user attachment: {filename}\n', file=f)

            except Exception as e:
                with open('validate_subs.log', 'a') as f:
                    print(f'Failed to attach file {att_path}: {str(e)}\n', file=f)
        elif att_path:
            with open('validate_subs.log', 'a') as f:
                print(f'User attachment file not found: {att_path}\n', file=f)

    text = message.as_string()

    smtp_user = os.getenv('SMTP_USER', email_config['smtp_user'])
    smtp_password = os.getenv('SMTP_PASSWORD')
    if not smtp_password:
        raise ValueError("SMTP_PASSWORD environment variable not set")

    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, customer, text)

    with open('validate_subs.log', 'a') as f:
        print(f'User email sent successfully to {customer}\n', file=f)

    return


def sendEmailToAdmin(subject, message, attachment_paths=None,
                     admin_email=None):
    """Send email to admin with optional file attachments."""
    try:
        with open('validate_subs.log', 'a') as f:
            print(f'Sending admin email: {subject}\n', file=f)

        # Load email configuration
        email_config, config_error = load_email_config()

        # Use admin_email from config if not provided
        if admin_email is None:
            admin_email = email_config['admin_email']

        # If config failed, send emergency notification
        if config_error:
            send_emergency_notification(
                config_error,
                admin_email,
                f"Email type: Admin notification\nRecipient: {admin_email}\nSubject: {subject}"
            )

        port = 465  # For SSL
        context = ssl.create_default_context()

        email_message = MIMEMultipart()
        email_message["From"] = email_config['from_address']
        email_message["To"] = admin_email
        email_message["Subject"] = subject
        # Add BCC recipients from config
        email_message["Bcc"] = ', '.join(email_config['bcc_list'])

        with open('validate_subs.log', 'a') as f:
            print(f'[SEND ADMIN EMAIL] BCC list: {", ".join(email_config["bcc_list"])}\n', file=f)

        # Add body to email
        email_message.attach(MIMEText(message, "plain"))

        # Add file attachments if provided
        if attachment_paths:
            for file_path in attachment_paths:
                if os.path.exists(file_path):
                    try:
                        with open(file_path, "rb") as attachment:
                            part = MIMEBase("application", "octet-stream")
                            part.set_payload(attachment.read())

                        # Encode file in ASCII characters to send by email
                        encoders.encode_base64(part)

                        # Add header as key/value pair to attachment part
                        filename = os.path.basename(file_path)
                        part.add_header(
                            "Content-Disposition",
                            f"attachment; filename= {filename}",
                        )

                        # Add attachment to message
                        email_message.attach(part)

                        with open('validate_subs.log', 'a') as f:
                            print(f'Added attachment: {filename}\n', file=f)

                    except Exception as e:
                        with open('validate_subs.log', 'a') as f:
                            print(f'Failed to attach file {file_path}: {str(e)}\n', file=f)
                else:
                    with open('validate_subs.log', 'a') as f:
                        print(
                            f'Attachment file not found: {file_path}\n', file=f)

        # Convert message to string and send
        text = email_message.as_string()

        smtp_user = os.getenv('SMTP_USER', email_config['smtp_user'])
        smtp_password = os.getenv('SMTP_PASSWORD')
        if not smtp_password:
            raise ValueError("SMTP_PASSWORD environment variable not set")

        with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, admin_email, text)

        with open('validate_subs.log', 'a') as f:
            print(f'Admin email sent successfully to {admin_email}\n', file=f)

    except Exception as e:
        with open('validate_subs.log', 'a') as f:
            print(f'Failed to send admin email: {str(e)}\n', file=f)
        # Don't raise exception - email failure shouldn't break the main
        # process


def call_code_a_validation(org_id, period, subscriber_file_path):
    """
    Call Code A validation subprocess and handle results.

    Args:
        org_id (str): Organization ID
        period (str): Filing period
        subscriber_file_path (str): Full path to subscriber CSV file

    Returns:
        dict: {
            'status': 'valid'|'invalid'|'error',
            'return_code': int,
            'csv_path': str or None,
            'excel_path': str or None,
            'artifact_paths': list,
            'error_message': str or None,
            'stdout': str,
            'stderr': str
        }
    """

    # Build Code A command
    base_output_dir = f"/var/www/broadband/uploads/{org_id}/{period}"
    code_a_base_dir = "/var/www/broadband"

    cmd = [
        "python3",
        "-m",
        "src.main",
        subscriber_file_path,
        str(org_id),
        period
    ]

    with open('validate_subs.log', 'a') as f:
        print(f'Calling Code A validation from {code_a_base_dir}: {" ".join(cmd)}\n', file=f)

    try:
        # Execute Code A subprocess with correct working directory
        result = subprocess.run(
            cmd,
            cwd=code_a_base_dir,
            capture_output=True,
            text=True,
            timeout=600  # 10 minute timeout
        )

        return_code = result.returncode
        stdout = result.stdout
        stderr = result.stderr

        with open('validate_subs.log', 'a') as f:
            print(
                f'Code A completed with return code: {return_code}\n',
                file=f)
            if stdout:
                print(f'Code A stdout: {stdout}\n', file=f)
            if stderr:
                print(f'Code A stderr: {stderr}\n', file=f)

        # Give filesystem time to sync files to disk (Code A writes CSV/Excel files)
        # This ensures files are fully written before Code B tries to read them
        with open('validate_subs.log', 'a') as f:
            print(f'Waiting 2 seconds for filesystem sync...\n', file=f)
        sleep(2)

        # Find all artifacts created by Code A
        # Code A saves files to new directory structure: /var/www/broadband/Subscriber_File_Validations/{period}/{org_id}/
        validation_results_dir = os.path.join("/var/www/broadband/Subscriber_File_Validations", period, str(org_id))
        artifact_paths = []

        if os.path.exists(validation_results_dir):
            # Get all files in company_id directory
            artifact_paths = glob.glob(f"{validation_results_dir}/*")
            with open('validate_subs.log', 'a') as f:
                print(f'Found {len(artifact_paths)} Code A artifacts in {validation_results_dir}\n',file=f)
                for path in artifact_paths:
                    print(f'  - {os.path.basename(path)}\n', file=f)
        else:
            with open('validate_subs.log', 'a') as f:
                print(f'Warning: Code A output directory not found: {validation_results_dir}\n', file=f)

        # Determine file paths for key outputs
        csv_path = None
        excel_path = None
        original_csv_path = None

        with open('validate_subs.log', 'a') as f:
            print(f'Searching for CSV/Excel files in artifacts:\n', file=f)
            for path in artifact_paths:
                print(f'  Checking: {path}\n', file=f)

        for path in artifact_paths:
            filename = os.path.basename(path)
            if filename.endswith('_Corrected_Subscribers.csv'):
                csv_path = path
                with open('validate_subs.log', 'a') as f:
                    print(f'Found CSV: {csv_path}\n', file=f)
            elif filename.endswith('_Corrected_Subscribers.xlsx'):
                excel_path = path
                with open('validate_subs.log', 'a') as f:
                    print(f'Found Excel: {excel_path}\n', file=f)
            elif filename.endswith('_Column_Count_Errors.xlsx'):
                # Column count error file takes precedence (it means validation couldn't even start)
                excel_path = path
                with open('validate_subs.log', 'a') as f:
                    print(f'Found Column Count Error Excel: {excel_path}\n', file=f)
            elif filename.endswith('_Original.csv') or (filename.endswith('.csv') and '_cleaned_temp' not in filename and '_Corrected_Subscribers' not in filename):
                # Find the original CSV file (ends with _Original.csv or is a CSV that's not a temp/corrected file)
                original_csv_path = path
                with open('validate_subs.log', 'a') as f:
                    print(f'Found Original CSV: {original_csv_path}\n', file=f)

        if not csv_path:
            with open('validate_subs.log', 'a') as f:
                print(f'WARNING: No CSV file found matching pattern *_Corrected_Subscribers.csv\n', file=f)
                print(f'Available files: {[os.path.basename(p) for p in artifact_paths]}\n', file=f)

        # Interpret return code
        if return_code == 0:
            status = 'valid'
            error_message = None
        elif return_code == 1:
            status = 'invalid'
            error_message = "File requires manual review before submission"
        elif return_code == 2:
            status = 'header_error'
            error_message = "Column headers do not match required format"
        else:
            status = 'error'
            error_message = f"Code A validation failed with return code {return_code}"
            if stderr:
                error_message += f": {stderr}"

        return {
            'status': status,
            'return_code': return_code,
            'csv_path': csv_path,
            'excel_path': excel_path,
            'original_csv_path': original_csv_path,
            'artifact_paths': artifact_paths,
            'error_message': error_message,
            'stdout': stdout,
            'stderr': stderr
        }

    except subprocess.TimeoutExpired:
        error_msg = "Code A validation timed out after 10 minutes"
        with open('validate_subs.log', 'a') as f:
            print(f'Code A timeout error: {error_msg}\n', file=f)

        return {
            'status': 'error',
            'return_code': 124,  # Standard timeout exit code
            'csv_path': None,
            'excel_path': None,
            'original_csv_path': None,
            'artifact_paths': [],
            'error_message': error_msg,
            'stdout': '',
            'stderr': error_msg
        }

    except Exception as e:
        error_msg = f"Failed to execute Code A validation: {str(e)}"
        with open('validate_subs.log', 'a') as f:
            print(f'Code A execution error: {error_msg}\n', file=f)

        return {
            'status': 'error',
            'return_code': 999,  # Custom error code
            'csv_path': None,
            'excel_path': None,
            'original_csv_path': None,
            'artifact_paths': [],
            'error_message': error_msg,
            'stdout': '',
            'stderr': str(e)
        }


def create_subscription(subfile, filename, isp, periodpath, period, user_email):
    print(
        "subfile ",
        subfile,
        " filename ",
        filename,
        " isp ",
        isp,
        " periodpath ",
        periodpath,
        " user_email ",
        user_email)
    subscrfile = periodpath + "/subscribers/" + subfile

    # === NEW: PHASE 1 - CODE A VALIDATION ===
    print("========================================")
    print("PHASE 1: Calling Code A for validation")
    print("========================================")

    # Call Code A validation
    validation_result = call_code_a_validation(isp, period, subscrfile)

    # Email all artifacts to admin regardless of outcome
    subject = f"Code A Validation Results - Org {isp}, Period {period}"
    if validation_result['status'] == 'valid':
        message = f"Code A validation completed successfully for Org {isp}.\n\nFile Status: VALID - Ready for geocoding and processing.\n\nReturn Code: {validation_result['return_code']}\n\nProcessed File: {validation_result['csv_path']}"
    elif validation_result['status'] == 'invalid':
        # Log stdout content for debugging
        with open('validate_subs.log', 'a') as f:
            print(f"[DEBUG EMAIL] Building admin email for invalid result\n", file=f)
            print(f"[DEBUG EMAIL] stdout length = {len(validation_result['stdout'])} characters\n", file=f)
            print(f"[DEBUG EMAIL] stdout content preview (first 500 chars):\n{validation_result['stdout'][:500]}\n", file=f)
            print(f"[DEBUG EMAIL] stderr length = {len(validation_result['stderr'])} characters\n", file=f)

        # Include full stdout for debugging why validation failed
        message = f"Code A validation completed for Org {isp}.\n\nFile Status: INVALID - Requires manual review.\n\nReason: {validation_result['error_message']}\n\nReturn Code: {validation_result['return_code']}\n\n{'='*60}\nDEBUG OUTPUT (stdout):\n{'='*60}\n{validation_result['stdout']}\n\n{'='*60}\nERROR OUTPUT (stderr):\n{'='*60}\n{validation_result['stderr']}\n\nCorrected file has been sent to user for review."
    elif validation_result['status'] == 'header_error':
        message = f"Code A validation FAILED for Org {isp}.\n\nFile Status: HEADER ERROR - Column headers do not match required format.\n\nReturn Code: {validation_result['return_code']}\n\nHeader-specific email has been sent to user."
    else:  # error
        message = f"Code A validation FAILED for Org {isp}.\n\nFile Status: ERROR\n\nError: {validation_result['error_message']}\n\nReturn Code: {validation_result['return_code']}\n\n{'='*60}\nDEBUG OUTPUT (stdout):\n{'='*60}\n{validation_result['stdout']}\n\n{'='*60}\nERROR OUTPUT (stderr):\n{'='*60}\n{validation_result['stderr']}"

    sendEmailToAdmin(subject, message, validation_result['artifact_paths'])

    # Handle validation results
    if validation_result['status'] == 'invalid':
        # File needs manual review - send corrected Excel to user and stop
        # processing
        print("========================================")
        print("PHASE 1 RESULT: INVALID - Sending corrected file to user")
        print("========================================")

        # Use provided user email
        customer = user_email
        cname = ''  # Get name from database for personalization
        with open('validate_subs.log', 'a') as f:
            print(f'[INVALID FILE] Using provided email: {customer} for org_id={isp}\n', file=f)

        # Try to get user name from database for personalization
        try:
            sql = """Select name from broadband.users where org_id = """ + isp + """ limit 1"""
            ps_cursor.execute(sql)
            userems = ps_cursor.fetchall()
            for em in userems:
                cname = em["name"]
            if cname:
                with open('validate_subs.log', 'a') as f:
                    print(f'[INVALID FILE] Found user name: {cname}\n', file=f)
            else:
                cname = 'Customer'  # Default if name not found
        except Exception as e:
            with open('validate_subs.log', 'a') as f:
                print(f'[INVALID FILE] Could not retrieve name from database: {e}\n', file=f)
            cname = 'Customer'  # Default if lookup fails

        # Create user message
        user_message = f"""Dear {cname},

Thank you for uploading your subscriber file to Regulatory Solutions for FCC BDC processing.

Initial review of the file suggests the file needs a little help.

We have attached a file called {isp}_Corrected_Subscribers.xlsx to this email.
This file is your original subscriber file that has been partially corrected.
This file identifies errors that need to be manually corrected. The file has color-coded cells:
- Green cells have been automatically corrected to match USPS standards
- Red and Pink cells need manual correction, and
- Yellow cells should be reviewed and corrected as appropriate (e.g., missing zip codes)

Please:
1. Open the attached {isp}_Modified_Subscribers.xlsx file
2. Correct the Red and Pink data cells
3. Review and correct all Yellow cells, if you can easily otherwise the next pass will autocorrect these cells
4. Complete your repairs and SAVE THE FILE IN A CSV format, then
5. Re-upload the corrected CSV file

If the file passes inspection, you will receive an email with a complete validation report.

For detailed requirements and instructions, please refer to:
https://regulatorysolutions.us/downloads/subscriber_template_instructionsV2.pdf

Thanks for taking care of this.
If you need help, please contact RSI at 972-836-7107.

Best regards,
The Regulatory Solutions Team"""

        # Process Excel file: remove OrigRowNum column and save with new name
        excel_attachment = validation_result['excel_path'] if validation_result['excel_path'] and os.path.exists(
            validation_result['excel_path']) else None

        if excel_attachment:
            try:
                import pandas as pd
                import openpyxl
                from openpyxl.styles import PatternFill

                # Read the original Excel file with all formatting
                wb = openpyxl.load_workbook(excel_attachment)
                ws = wb.active

                # Find OrigRowNum column index (should be first column)
                header_row = [cell.value for cell in ws[1]]
                if 'OrigRowNum' in header_row:
                    origrownum_col_idx = header_row.index('OrigRowNum') + 1  # openpyxl uses 1-based indexing

                    with open('validate_subs.log', 'a') as f:
                        print(f'Removing OrigRowNum column (column {origrownum_col_idx}) from Excel file\n', file=f)

                    # Delete the OrigRowNum column
                    ws.delete_cols(origrownum_col_idx)

                # Save with new filename
                excel_dir = os.path.dirname(excel_attachment)
                modified_excel_path = os.path.join(excel_dir, f'{isp}_modified_subscription_file.xlsx')
                wb.save(modified_excel_path)

                with open('validate_subs.log', 'a') as f:
                    print(f'Created modified Excel file: {os.path.basename(modified_excel_path)}\n', file=f)

                # Update attachment to use modified file
                excel_attachment = modified_excel_path

            except Exception as e:
                with open('validate_subs.log', 'a') as f:
                    print(f'Error processing Excel file: {str(e)}\n', file=f)
                    print(f'Sending original file instead\n', file=f)

        # Send only the corrected Excel file (not the original CSV)
        custom_subject = f'Your FCC BDC Subscriber File Failed to Complete Processing due to Errors; Action Requested ({isp})'
        sendEmail(
            customer,
            cname,
            user_message,
            excel_attachment,
            custom_subject)

        if excel_attachment:
            with open('validate_subs.log', 'a') as f:
                print(f'Sent corrected Excel file to user: {os.path.basename(excel_attachment)}\n',file=f)
        else:
            with open('validate_subs.log', 'a') as f:
                print(
                    f'Warning: No Excel file available to send to user for org {isp}\n',
                    file=f)

        # Update database status
        with open('validate_subs.log', 'a') as f:
            print(f'updating processing status and adding messages\n',file=f)

        sql = """Update filer_processing_status set subscription_processed = true, subscription_status = 'data_validation_failed' where org_id = """ + \
            isp + """ and filing_period = '""" + period + """' """
        cursor.execute(sql)
        conn.commit()

        sql =  f"""Insert into broadband.messages (message_type, message,datetime, org_id) values ('subscriber','Subscriber file processing error. Check your email for details.', now(),{isp})"""
        with open('validate_subs.log', 'a') as f:
            print(f'inserting message {sql}\n',file=f)
        cursor.execute(sql)
        conn.commit()

        return  # Stop processing

    elif validation_result['status'] == 'header_error':
        # Header validation error - send header-specific email
        print("========================================")
        print("PHASE 1 RESULT: HEADER ERROR - Sending header-specific email to user")
        print("========================================")

        # Use provided user email
        customer = user_email
        cname = ''  # Get name from database for personalization
        with open('validate_subs.log', 'a') as f:
            print(f'[HEADER ERROR] Using provided email: {customer} for org_id={isp}\n', file=f)

        # Try to get user name from database for personalization
        try:
            sql = """Select name from broadband.users where org_id = """ + isp + """ limit 1"""
            ps_cursor.execute(sql)
            userems = ps_cursor.fetchall()
            for em in userems:
                cname = em["name"]
            if cname:
                with open('validate_subs.log', 'a') as f:
                    print(f'[HEADER ERROR] Found user name: {cname}\n', file=f)
            else:
                cname = 'Customer'  # Default if name not found
        except Exception as e:
            with open('validate_subs.log', 'a') as f:
                print(f'[HEADER ERROR] Could not retrieve name from database: {e}\n', file=f)
            cname = 'Customer'  # Default if lookup fails

        header_error_message = f"""Dear {cname},

Thank you for uploading your subscriber file to Regulatory Solutions for FCC BDC processing.

We were unable to process your file because the column headers do not match the required format.

Your file must contain exactly these 12 column headers (in any order):
• customer
• lat
• lon
• address
• city
• state
• zip
• download
• upload
• voip_lines_quantity
• business_customer
• technology

Common Issues:
- Column headers have extra spaces or special characters
- Headers are in a different row (not the first row)
- Headers are misspelled or use different names
- File contains extra rows before the header row

What to do next:
1. Review your attached file and verify the column headers match exactly
2. Correct the headers to match the required names above
3. Ensure headers are in the first row of your file
4. Save your file and re-upload

For detailed requirements and a template, please refer to:
https://regulatorysolutions.us/downloads/subscriber_template_instructionsV2.pdf

If you need assistance, please contact RSI at 972-836-7107.

Best regards,
The Regulatory Solutions Team"""

        # Get original CSV to attach
        original_csv_attachment = validation_result.get('original_csv_path')
        if original_csv_attachment and os.path.exists(original_csv_attachment):
            with open('validate_subs.log', 'a') as f:
                print(f'Attaching original CSV to header error email: {original_csv_attachment}\n', file=f)
        else:
            original_csv_attachment = None

        header_email_subject = 'FCC BDC Subscriber File - Column Header Error'
        sendEmail(
            customer,
            cname,
            header_error_message,
            original_csv_attachment,
            header_email_subject)

        # Update database status
        with open('validate_subs.log', 'a') as f:
            print(f'updating processing status and adding messages\n',file=f)

        sql = """Update filer_processing_status set subscription_processed = true, subscription_status = 'header_validation_failed' where org_id = """ + \
            isp + """ and filing_period = '""" + period + """' """
        cursor.execute(sql)
        conn.commit()

        sql =  f"""Insert into broadband.messages (message_type, message,datetime, org_id) values ('subscriber','Subscriber file processing error. Check your email for details.', now(),{isp})"""
        with open('validate_subs.log', 'a') as f:
            print(f'inserting message {sql}\n',file=f)
        cursor.execute(sql)
        conn.commit()

        return  # Stop processing

    elif validation_result['status'] == 'error':
        # Code A failed - stop processing
        print("========================================")
        print("PHASE 1 RESULT: ERROR - Code A validation failed")
        print("========================================")

        # Use provided user email
        customer = user_email
        cname = ''  # Get name from database for personalization
        with open('validate_subs.log', 'a') as f:
            print(f'[VALIDATION ERROR] Using provided email: {customer} for org_id={isp}\n', file=f)

        # Try to get user name from database for personalization
        try:
            sql = """Select name from broadband.users where org_id = """ + isp + """ limit 1"""
            ps_cursor.execute(sql)
            userems = ps_cursor.fetchall()
            for em in userems:
                cname = em["name"]
            if cname:
                with open('validate_subs.log', 'a') as f:
                    print(f'[VALIDATION ERROR] Found user name: {cname}\n', file=f)
            else:
                cname = 'Customer'  # Default if name not found
        except Exception as e:
            with open('validate_subs.log', 'a') as f:
                print(f'[VALIDATION ERROR] Could not retrieve name from database: {e}\n', file=f)
            cname = 'Customer'  # Default if lookup fails

        # Check if this is a header validation error
        is_header_error = 'Could not locate valid column headers' in validation_result.get('error_message', '')

        # Create user error message based on error type
        if is_header_error:
            with open('validate_subs.log', 'a') as f:
                print(f'[VALIDATION ERROR] Header error detected - sending header-specific email\n', file=f)

            error_message = f"""Dear {cname},

Thank you for uploading your subscriber file to Regulatory Solutions for FCC BDC processing.

We were unable to process your file because the column headers do not match the required format.

Your file must contain exactly these 12 column headers (in any order):
• customer
• lat
• lon
• address
• city
• state
• zip
• download
• upload
• voip_lines_quantity
• business_customer
• technology

Common Issues:
- Column headers have extra spaces or special characters
- Headers are in a different row (not the first row)
- Headers are misspelled or use different names
- File contains extra rows before the header row

What to do next:
1. Review your attached file and verify the column headers match exactly
2. Correct the headers to match the required names above
3. Ensure headers are in the first row of your file
4. Save your file and re-upload

For detailed requirements and a template, please refer to:
https://regulatorysolutions.us/downloads/subscriber_template_instructionsV2.pdf

If you need assistance, please contact RSI at 972-836-7107.

Best regards,
The Regulatory Solutions Team"""

            email_subject = 'FCC BDC Subscriber File - Column Header Error'

        else:
            with open('validate_subs.log', 'a') as f:
                print(f'[VALIDATION ERROR] Generic error - sending standard error email\n', file=f)

            error_message = f"""Dear {cname},

We encountered a technical issue while processing your subscriber file that prevented validation from completing.

This is typically due to file format issues such as:
- Missing required columns
- Incorrect file structure
- File encoding problems

Please verify your file follows the subscriber template format and re-upload, or contact our support team for assistance.

For your convenience, detailed field requirements are available at:
https://regulatorysolutions.us/downloads/subscriber_template_instructionsV2.pdf

Best regards,
The Regulatory Solutions Team"""

            email_subject = 'Subscriber File Processing Error'

        # Send error notification to user with original CSV attached
        original_csv_attachment = validation_result.get('original_csv_path')
        if original_csv_attachment and os.path.exists(original_csv_attachment):
            with open('validate_subs.log', 'a') as f:
                print(f'Attaching original CSV to error email: {original_csv_attachment}\n', file=f)
        else:
            original_csv_attachment = None
            with open('validate_subs.log', 'a') as f:
                print(f'Original CSV not found for attachment\n', file=f)

        sendEmail(customer, cname, error_message, original_csv_attachment,
                  email_subject)

        # Update database status

        with open('validate_subs.log', 'a') as f:
            print(f'updating processing status and adding messages\n',file=f)

        sql =  """Insert into broadband.messages (message_type, message,datetime, org_id) values ('subscriber','Subscriber file processing error. Check your email for details.', now(),""" + isp + """ )"""
        with open('validate_subs.log', 'a') as f:
            print(f'inserting message {sql}\n',file=f)
        try:
            if cursor:
                cursor.execute(sql)
                conn.commit()
                with open('validate_subs.log', 'a') as f:
                    print(f'message inserted\n',file=f)
            else:
                with open('validate_subs.log', 'a') as f:
                    print(f'no cursor\n',file=f)
        except (Exception, psycopg2.DatabaseError) as error:
            with open('validate_subs.log', 'a') as f:
                print(f'error inserting data {error}\n',file=f)

        sql = """Update filer_processing_status set subscription_processed = true, subscription_status = 'system_error' where org_id = """ + \
            isp + """ and filing_period = '""" + period + """' """
        with open('validate_subs.log', 'a') as f:
            print(f'inserting this message {sql}\n',file=f)
        cursor.execute(sql)
        conn.commit()
        with open('validate_subs.log', 'a') as f:
            print(f'done updating filer_processing_status\n',file=f)

        return  # Stop processing

    # If we get here, validation_result['status'] == 'valid'
    print("========================================")
    print("PHASE 1 RESULT: VALID - Proceeding with geocoding and processing")
    print("========================================")

    # Use Code A's corrected CSV file for further processing
    if not validation_result['csv_path'] or not os.path.exists(
            validation_result['csv_path']):
        print("ERROR: Code A corrected CSV file not found")
        return

    subscrfile = validation_result['csv_path']  # Use Code A's output
    print(f"Using Code A corrected file: {subscrfile}")

    # === PHASE 2: CODE B GEOCODING AND DATABASE PROCESSING ===
    print("========================================")
    print("PHASE 2: Starting geocoding and database processing")
    print("========================================")

    # Continue with Code B's database operations (geocoding, tract assignment,
    # etc.)
    now = datetime.now()
    date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
    addrerr = []  # Only geocoding errors now

    with open(subscrfile) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        next(csv_reader)

        sql = """SELECT EXISTS (
                        SELECT 1
                        FROM pg_tables
                        WHERE schemaname = 'subscibers'
                        AND tablename = 'subs_""" + str(isp) + """'
                        );"""
        cursor.execute(sql)
        se = cursor.fetchone()
        subsexist = se[0]
        print("subsexist" + str(subsexist))
        if (subsexist == True):
            print("getting leads from subs")
            sql = """Select * into subscribers.subs_""" + \
                str(isp) + """_temp from subscribers.subs_""" + \
                str(isp) + """ where type != 'Active' """
            cursor.execute(sql)

        sql = """Drop table if exists subscribers.subs_""" + isp
        cursor.execute(sql)
        sql = """CREATE TABLE subscribers.subs_""" + isp + \
            """ (customer text,lat numeric,lon numeric,address text,address2 text,city text,state text,zip text,download numeric,upload numeric,voip_lines_quantity integer,business_customer numeric,technology integer,tech text,tract text, match boolean,bdc_id integer,type text, date timestamp without time zone,notes text)"""
        cursor.execute(sql)
        conn.commit()
        subsarr = []
        state = ''
        voipneeded = False
        line_count = 0

        for row in csv_reader:
            customer = row[0]
            lat = row[1]
            lon = row[2]
            address = row[3].upper()
            city = row[4].upper()
            state = row[5].upper()
            zip = row[6]
            down = row[7]
            up = row[8]
            voip = row[9]
            if voip == '':
                voip = int(0)
                print('voip 0')
            if int(voip) > 0:
                voipneeded = True
            business = row[10]
            techname = row[11]
            print('techname ', techname)

            # Technology code mapping (keep this since it's Code B specific)
            tech = 1
            if techname == 'wireless_unlicensed':
                tech = 70
            if techname == 'wireless_gaa':
                tech = 72
            if techname == 'wireless_pal':
                tech = 71
            if techname == 'wireless_educational':
                tech = 71
            if techname == 'fiber':
                tech = 50
            if techname == 'cable':
                tech = 43
            if techname == 'ethernet':
                tech = 10
            if techname == 'adsl2':
                tech = 11
            if techname == 'voip':
                tech = 1

            print(lat, lon, address, city, state, zip, str(tech))

            # GEOCODING: Only geocode if lat/lon are missing (Code A validated
            # addresses)
            if lat == '' or lon == '' and (
                    address != '' and city != '' and state != '' and zip != ''):
                addr = address + ',' + city + ',' + state + ' ' + zip
                (lt, ln) = geoCode(addr)
                if lt is None or ln is None:
                    rowerr = line_count + 1
                    print("error geocoding row " + str(rowerr))
                    errstr = 'error geocoding row ' + \
                        str(rowerr) + ' addr: ' + addr
                    addrerr.append(errstr)
                else:
                    lat = lt
                    lon = ln

            # Census tract assignment (only if we have coordinates)
            if lat != '' and lon != '':
                print("lat/lon" + str(lat) + " " + str(lon))
                statefp = 0
                if state != 'VI':
                    sql = """select statefp10 from census_data.states where ST_Intersects(geom,ST_setSRID(ST_Makepoint(%s,%s),4326))"""
                    cursor.execute(sql, (lon, lat))
                    print(sql, lon, lat)
                    sfp = cursor.fetchone()
                    print(sfp)
                    statefp = sfp[0]
                else:
                    statefp = '78'
                print("statefp " + statefp)
                sql = """Select geoid from census_data.tracts20 where statefp = %s and ST_Intersects(ST_SetSRID(ST_MakePoint(%s,%s),4326), geog)"""
                print(sql, lat, lon)
                cursor.execute(sql, (statefp, float(lon), float(lat)))
                t = cursor.fetchone()
                tract = t[0]
                print("tract " + tract)
                sql = """Insert into subscribers.subs_""" + isp + """ (customer,lat,lon,address,city,state,zip,download,upload,voip_lines_quantity,business_customer,technology,tech,tract,type,date)
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
                cursor.execute(
                    sql,
                    (customer,
                     lat,
                     lon,
                     address,
                     city,
                     state,
                     zip,
                     down,
                     up,
                     voip,
                     business,
                     tech,
                     techname,
                     tract,
                     'Active',
                     date_time))
                conn.commit()
            line_count += 1
            print({line_count}, end='\r')
        print(f'Processed {line_count} lines.')
        conn.commit()
        print("creating index on subs table ")
        sql = """create index subs_""" + \
            str(isp) + """_customer_index on subscribers.subs_""" + \
            str(isp) + """ (customer);"""
        cursor.execute(sql)
        conn.commit()

        # Handle geocoding errors (only errors now, Code A handled validation)
        if len(addrerr) > 0:
            errf = periodpath + "/subscription_processed/processing_errors.txt"
            dirExist = os.path.exists(periodpath + "/subscription_processed/")
            if dirExist == False:
                os.umask(0)
                os.makedirs(
                    periodpath +
                    "/subscription_processed/",
                    mode=0o777)
            isExist = os.path.exists(errf)

            if (isExist):
                os.remove(errf)
            errfil = open(
                periodpath +
                "/subscription_processed/processing_errors.txt",
                "a")
            errfil.write('Date: ' + date_time + '\n')

            # Use provided user email
            customer = user_email
            cname = ''  # Get name from database for personalization
            with open('validate_subs.log', 'a') as f:
                print(f'[GEOCODING ERRORS] Using provided email: {customer} for org_id={isp}\n', file=f)

            # Try to get user name from database for personalization
            try:
                sql = """Select name from broadband.users where org_id = """ + isp + """ limit 1"""
                ps_cursor.execute(sql)
                userems = ps_cursor.fetchall()
                for em in userems:
                    cname = em["name"]
                if cname:
                    with open('validate_subs.log', 'a') as f:
                        print(f'[GEOCODING ERRORS] Found user name: {cname}\n', file=f)
                else:
                    cname = 'Customer'  # Default if name not found
            except Exception as e:
                with open('validate_subs.log', 'a') as f:
                    print(f'[GEOCODING ERRORS] Could not retrieve name from database: {e}\n', file=f)
                cname = 'Customer'  # Default if lookup fails

            em_message = 'Dear ' + cname + \
                ', \nYour subscriber file passed validation but we encountered geocoding errors for some addresses:\n\nDate: ' + date_time + '\n'

            for errs in addrerr:
                errfil.write(errs + "\n")
                em_message += errs + "\n"

            em_message += '\nThese addresses could not be geocoded and may need manual coordinate entry. The rest of your file has been processed successfully.\n\nBest Regards,\n\nThe Regulatory Solutions Team'
            print(
                "sending geocoding error email to  " +
                customer +
                " " +
                cname)
            sendEmail(customer, cname, em_message, None,
                      'Subscriber File Processing - Geocoding Issues')
            errfil.close()

            with open('validate_subs.log', 'a') as f:
                print(f'updating processing status and adding messages\n',file=f)

            sql = """Update filer_processing_status set subscription_processed = true, subscription_status = 'geocoding_errors' where org_id = """ + \
                isp + """ and filing_period = '""" + period + """' """
            cursor.execute(sql)
            conn.commit()

            sql =  f"""Insert into broadband.messages (message_type, message,datetime, org_id) values ('subscriber','Subscriber file processing error. Check your email for details.', now(),{isp})"""
            with open('validate_subs.log', 'a') as f:
                print(f'inserting message {sql}\n',file=f)
            cursor.execute(sql)
            conn.commit()

        else:
            # SUCCESS - continue with existing Code B processing
            if (subsexist == True):
                sql = """insert into subscribers.subs_""" + isp + \
                    """ select * from subscribers.subs_""" + isp + """_temp """
                cursor.execute(sql)
                conn.commit()

            sql = """Drop table if exists subscribers.subs_""" + isp + """_temp """
            cursor.execute(sql)
            conn.commit()

            # Continue with successful processing - create output files
            tmpout = "/tmp/" + isp + "_subscription_processed.csv"
            outfil = periodpath + "/subscription_processed/" + \
                isp + "_subscription_processed.csv"
            dirExist = os.path.exists(periodpath + "/subscription_processed/")
            if dirExist == False:
                os.umask(0)
                os.makedirs(
                    periodpath +
                    "/subscription_processed/",
                    mode=0o777)
            isExist = os.path.exists(outfil)

            if (isExist):
                os.remove(outfil)

            sql = """COPY (Select tract,
                            technology,
                            download,
                            upload,
                            count(customer) as total,
                            count(customer) - sum(business_customer) as residential
                            from subscribers.subs_""" + isp + """ where technology > 1 and type = 'Active'
                            group by tract,technology,download,upload
                            order by tract, download, upload) to '""" + tmpout + """' with CSV DELIMITER ','  """
            cursor.execute(sql)
            fcnt = 1
            with open(tmpout, 'r') as f:
                lines = f.readlines()
                lenl = len(lines)
                with open(outfil, 'w') as fo:
                    for l in lines:
                        if "\n" in l:
                            print("found new line")
                            nlpos = l.find("\n")
                            nl = l[:nlpos]
                            print("nl", nl)
                            if fcnt < lenl:
                                nl = nl + "\n"
                            fo.write(nl)
                            fcnt += 1

            # Create 477 version (change 71 to 70)
            tmpout = "/tmp/477_" + isp + "_subscription_processed.csv"
            outfil = periodpath + "/subscription_processed/477_" + \
                isp + "_subscription_processed.csv"
            sql = """COPY (Select tract,
                            case when technology = 71 then 70
                            else technology
                            end as techcode,
                            download,
                            upload,
                            count(customer) as total,
                            count(customer) - sum(business_customer) as residential
                            from subscribers.subs_""" + isp + """ where technology > 1
                             group by tract,techcode,download,upload) to '""" + tmpout + """' with CSV DELIMITER ','  """
            cursor.execute(sql)
            with open(tmpout) as f:
                contents = f.read()
            with open(outfil, "w") as fo:
                print(contents, file=fo)

            # Handle VoIP processing if needed
            if voipneeded == True:
                tmpout = "/tmp/" + isp + "voice_subscription_processed.csv"
                outfil = periodpath + "/subscription_processed/" + \
                    isp + "_voice_subscription_processed.csv"
                isExist = os.path.exists(outfil)
                if (isExist):
                    os.remove(outfil)

                sql = """COPY (Select tract,
                        '1' as service_type,
                        sum(voip_lines_quantity) as total,
                        sum(voip_lines_quantity) - (sum(business_customer * voip_lines_quantity)) as residential
                        from subscribers.subs_""" + isp + """ where voip_lines_quantity > 0 group by tract order by tract) to '""" + tmpout + """' with CSV DELIMITER ','  """
                cursor.execute(sql)
                fcnt = 1
                with open(tmpout, 'r') as f:
                    lines = f.readlines()
                    lenl = len(lines)
                    with open(outfil, 'w') as fo:
                        for l in lines:
                            if "\n" in l:
                                print("found new line")
                                nlpos = l.find("\n")
                                nl = l[:nlpos]
                                print("nl", nl)
                                if fcnt < lenl:
                                    nl = nl + "\n"
                                fo.write(nl)
                                fcnt += 1

                # Create voice state data
                outfil = periodpath + "/subscription_processed/" + isp + "_voice_state_data.txt"
                sleep(5)
                sql = 'select distinct substring(tract,1,2) as statefips from subscribers.subs_' + str(
                    isp)
                print(sql)
                ps_cursor.execute(sql)
                states = ps_cursor.fetchall()
                for state in states:
                    contents = ''
                    contents = "state " + str(state["statefips"]) + "\n"
                    print("adding voip for state " + str(state["statefips"]))
                    sql = """Select
                            case when technology >= 71 then 70
                            else technology
                            end as techcode,
                            sum(voip_lines_quantity) as total
                            from subscribers.subs_""" + isp + """ where voip_lines_quantity > 0 and substring(tract,1,2) = '""" + state["statefips"] + """' group by techcode  """
                    cursor.execute(sql)
                    techsums = cursor.fetchall()
                    if techsums is not None:
                        for t in techsums:
                            contents += "tech code " + \
                                str(t[0]) + ": " + str(t[1]) + "\n"
                        with open(outfil, "a") as ts:
                            print(contents, file=ts)

            # Update final status to complete
            sql = """Update filer_processing_status set subscription_processed = true, subscription_status = 'complete' where org_id = """ + \
                isp + """ and filing_period = '""" + period + """' """
            cursor.execute(sql)
            conn.commit()

            sql =  f"""Insert into broadband.messages (message_type, message,datetime, org_id) values ('subscriber','Subscriber file processing complete', now(),{isp})"""
            with open('validate_subs.log', 'a') as f:
                print(f'inserting message {sql}\n',file=f)
            cursor.execute(sql)
            conn.commit()


            # Send Phase 2 completion email to admin
            print("========================================")
            print("PHASE 2 COMPLETE: Sending results to admin")
            print("========================================")

            phase2_subject = f"Code B Processing Results - Org {isp}, Period {period}"

            # Collect Phase 2 output files
            subscription_dir = periodpath + "/subscription_processed/"
            phase2_files = []
            if os.path.exists(subscription_dir):
                phase2_files = glob.glob(f"{subscription_dir}/*")
                with open('validate_subs.log', 'a') as f:
                    print(f'Found {len(phase2_files)} Phase 2 output files\n', file=f)
                    for filepath in phase2_files:
                        print(f'  - {os.path.basename(filepath)}\n', file=f)

            # Build Phase 2 completion message
            phase2_message = f"""Code B processing completed successfully for Org {isp}.

File Status: COMPLETE - All geocoding and database operations finished.

Period: {period}
Final Status: complete
Total Rows Processed: {line_count}
Geocoding Errors: {len(addrerr)}
VoIP Lines Included: {'Yes' if voipneeded else 'No'}

Output Files Created:
"""
            if phase2_files:
                for filepath in phase2_files:
                    file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
                    phase2_message += f"  - {os.path.basename(filepath)} ({file_size} bytes)\n"
            else:
                phase2_message += "  (No output files found)\n"

            phase2_message += f"""
Output Directory: {subscription_dir}

Database Table: subscribers.subs_{isp}

Processing completed at: {datetime.now().strftime("%m/%d/%Y, %H:%M:%S")}

All files are attached for manual inspection.
"""

            sendEmailToAdmin(phase2_subject, phase2_message, phase2_files)

            with open('validate_subs.log', 'a') as f:
                print(f'Phase 2 completion email sent to admin with {len(phase2_files)} attachments\n', file=f)

            # Use provided user email
            customer = user_email
            cname = ''  # Get name from database for personalization
            with open('validate_subs.log', 'a') as f:
                print(f'[PHASE 2 SUCCESS] Using provided email: {customer} for org_id={isp}\n', file=f)

            # Try to get user name from database for personalization
            try:
                sql = """Select name from broadband.users where org_id = """ + isp + """ limit 1"""
                ps_cursor.execute(sql)
                userems = ps_cursor.fetchall()
                for em in userems:
                    cname = em["name"]
                if cname:
                    with open('validate_subs.log', 'a') as f:
                        print(f'[PHASE 2 SUCCESS] Found user name: {cname}\n', file=f)
                else:
                    cname = 'Customer'  # Default if name not found
            except Exception as e:
                with open('validate_subs.log', 'a') as f:
                    print(f'[PHASE 2 SUCCESS] Could not retrieve name from database: {e}\n', file=f)
                cname = 'Customer'  # Default if lookup fails

            if customer:
                with open('validate_subs.log', 'a') as f:
                    print(f'[PHASE 2 SUCCESS] Sending success email to: {cname} <{customer}> for org_id={isp}\n', file=f)

                # Find the VR.xlsx file to attach
                # VR file is created by Code A in Subscriber_File_Validations directory
                vr_file = None

                # First try to find in phase2_files (subscription_processed directory)
                for filepath in phase2_files:
                    if filepath.endswith('_VR.xlsx'):
                        vr_file = filepath
                        with open('validate_subs.log', 'a') as f:
                            print(f'[PHASE 2 SUCCESS] Found VR file in phase2_files: {os.path.basename(vr_file)}\n', file=f)
                        break

                # If not found, search in Subscriber_File_Validations directory where Code A creates it
                if not vr_file:
                    validation_dir = f"/var/www/broadband/Subscriber_File_Validations/{period}/{isp}"
                    with open('validate_subs.log', 'a') as f:
                        print(f'[PHASE 2 SUCCESS] Searching for VR file in validation directory: {validation_dir}\n', file=f)

                    if os.path.exists(validation_dir):
                        vr_files = glob.glob(f"{validation_dir}/*_VR.xlsx")
                        if vr_files:
                            vr_file = vr_files[0]  # Take the first match
                            with open('validate_subs.log', 'a') as f:
                                print(f'[PHASE 2 SUCCESS] Found VR file in validation directory: {os.path.basename(vr_file)}\n', file=f)
                        else:
                            with open('validate_subs.log', 'a') as f:
                                print(f'[PHASE 2 SUCCESS] WARNING: No VR.xlsx files found in {validation_dir}\n', file=f)
                    else:
                        with open('validate_subs.log', 'a') as f:
                            print(f'[PHASE 2 SUCCESS] WARNING: Validation directory does not exist: {validation_dir}\n', file=f)

                if not vr_file:
                    with open('validate_subs.log', 'a') as f:
                        print(f'[PHASE 2 SUCCESS] WARNING: VR.xlsx file not found in any location\n', file=f)

                # Create success message for user
                success_message = f"""Dear {cname},

Thank you for submitting your subscriber file to Regulatory Solutions for FCC BDC processing.

We are pleased to inform you that your subscriber file has been successfully processed and validated.

Your data has been geocoded, validated against census tract boundaries, and prepared for FCC submission.

Attached to this email is your complete Validation Report ({isp}_VR.xlsx). This report contains:
- All corrections made to your data (addresses, coordinates, formatting, etc.)
- Smarty API address validations performed
- Any duplicate records that were renamed
- A complete processing log

We recommend reviewing this report and updating your source database to reflect these corrections for future submissions.

If we discover any issues during our final review, we will contact you promptly. Otherwise, your subscriber file submission is complete and no further action is required on your part.

Thank you for your timely attention to this important filing requirement.

Best regards,
The Regulatory Solutions Team"""

                success_subject = f'FCC BDC Subscriber File Successfully Processed ({isp})'
                sendEmail(
                    customer,
                    cname,
                    success_message,
                    vr_file,  # Attach VR.xlsx file
                    success_subject)

                with open('validate_subs.log', 'a') as f:
                    if vr_file:
                        print(f'[PHASE 2 SUCCESS] Success email sent to user with VR attachment: {customer}\n', file=f)
                    else:
                        print(f'[PHASE 2 SUCCESS] Success email sent to user (no VR attachment found): {customer}\n', file=f)
            else:
                with open('validate_subs.log', 'a') as f:
                    print(f'[PHASE 2 SUCCESS] WARNING: No user found in database for org_id={isp}\n', file=f)

    return

def main(ispid,per,user_email):
	# Check command line arguments
	"""if len(sys.argv) < 4:
	    print("ERROR: Missing required arguments")
	    print("Usage: python3 validate_subscription_isp_mod_2.py {isp_id} yyyy-mm-dd {user_email}")
	    print("Example: python3 validate_subscription_isp_mod_2.py 123 2025-06-30 user@example.com")
	    sys.exit(1)

	ispid = sys.argv[1]
	per = sys.argv[2]
	user_email = sys.argv[3]
	"""
	ispid = str(ispid)
	# Validate email format (basic validation)
	email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
	if not re.match(email_pattern, user_email):
	    print(f"ERROR: Invalid email format: {user_email}")
	    print("Please provide a valid email address")
	    sys.exit(1)

	now = datetime.now()
	x = now.strftime("%m/%d/%Y, %H:%M:%S")

	with open('validate_subs.log', 'a') as f:
	    print(x, file=f)
	    print(f'validate_subscription_isp run for ISP {ispid}, Period {per}, User Email {user_email}\n', file=f)


	db_host = os.getenv('DB_HOST', 'localhost')
	db_port = os.getenv('DB_PORT', '5432')
	db_name = os.getenv('DB_NAME', 'broadband')
	db_user = os.getenv('DB_USER', 'broadband')
	db_password = os.getenv('DB_PASSWORD')
	if not db_password:
	    raise ValueError("DB_PASSWORD environment variable not set")

	global conn
	conn = psycopg2.connect(
	    database=db_name,
	    user=db_user,
	    password=db_password,
	    host=db_host,
	    port=db_port
	)
	global ps_cursor
	global cursor
	ps_cursor = conn.cursor(cursor_factory=RealDictCursor)
	cursor = conn.cursor()

	# Set initial subscription_status to 'processing' at start of validation
	try:
	    with open('validate_subs.log', 'a') as f:
	        print(f'Setting subscription_status to "processing" for org_id={ispid}, period={per}\n', file=f)

	    # Try UPDATE first
	    sql = """UPDATE broadband.filer_processing_status
	             SET subscription_processed = false, subscription_status = 'processing'
	             WHERE org_id = %s AND filing_period = %s"""
	    cursor.execute(sql, (ispid, per))

	    # If no rows were updated, INSERT new row
	    if cursor.rowcount == 0:
	        with open('validate_subs.log', 'a') as f:
	            print(f'No existing record found - inserting new row for org_id={ispid}, period={per}\n', file=f)

	        sql = """INSERT INTO broadband.filer_processing_status
	                 (org_id, filing_period, subscription_processed, subscription_status)
	                 VALUES (%s, %s, false, 'processing')"""
	        cursor.execute(sql, (ispid, per))

	    conn.commit()

	    with open('validate_subs.log', 'a') as f:
	        print(f'Successfully set subscription_status to "processing" for org_id={ispid}, period={per}\n', file=f)

	except Exception as e:
	    with open('validate_subs.log', 'a') as f:
	        print(f'WARNING: Could not set processing status for org_id={ispid}, period={per}: {e}\n', file=f)
	        print(f'Continuing with validation process...\n', file=f)
	    # Continue processing anyway - we'll still update status at the end

	# get all the wisps so we can check for aps done later
	# sql = """Select org_id from broadband.fixed_wireless_data where ap_data = true""";
	# ps_cursor.execute(sql);
	# completeisps = ps_cursor.fetchall()
	# ps_cursor.close()


	dir_path = r'/var/www/broadband/uploads'
	procisp = 0
	endperiod = ''
	for path in os.listdir(dir_path):
	    if os.path.isdir(os.path.join(dir_path, path)):
	        # at isp directory
	        print("isp ", os.path.basename(path))
	        procisp = os.path.basename(path)
                
	        if procisp == ispid:
	            print("found isp...processing");
	            isppth = dir_path + '/' + os.path.basename(path)
	            for isp in os.listdir(isppth):
	                if os.path.isdir(os.path.join(isppth, isp)):
	                    # at period directory
	                    endperiod = os.path.basename(isp)
	                    print("    period ", endperiod)
	                    print("    ", os.path.basename(isp))
	                    periodpath = isppth + '/' + os.path.basename(isp)
	                    for period in os.listdir(periodpath):
	                        if os.path.isdir(os.path.join(periodpath, period)):
	                            print("       ", os.path.basename(period))
	                            # sleep(3)
	                            dirname = os.path.basename(period)

	                            if dirname == 'subscribers' and endperiod == per:
	                                print(
	                                    "          building subscription file from subscribers for isp ", procisp)
	                                # go build subscription file
	                                subpath = periodpath + '/' + \
	                                    os.path.basename(period)
	                                print("subpath", subpath)
	                                for subfile in os.listdir(subpath):
	                                    print("subfile", subfile)
	                                    if os.path.isfile(
	                                            os.path.join(subpath, subfile)):
	                                        sfile = os.path.basename(subfile)
	                                        print("sfile", sfile)
	                                        print(
	                                            "          processing subscribers file ", subfile)
	                                        create_subscription(
	                                            subfile, sfile, procisp, periodpath, endperiod, user_email)

	conn.close()

	with open('validate_subs.log', 'a') as f:
	    print('validate_subscription_isp run done for ' + str(ispid) + '\n', file=f)
 
