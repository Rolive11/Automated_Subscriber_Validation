"""File handling utilities for reading, writing, and managing files."""

import os
import sys
import re
import shutil
import csv
import pandas as pd
import numpy as np
from datetime import datetime
import time
import json
import traceback
import openpyxl
from openpyxl.styles import PatternFill
from src.utils.logging import debug_print
from src.config.settings import EXPECTED_COLUMNS, VALID_STATES, VALID_TECHNOLOGIES, STATE_LAT_RANGES, STATE_LON_RANGES, DTYPE_DICT, GREEN_FILL, PINK_FILL, YELLOW_FILL, RED_FILL
from src.validation.customer import validate_customer_uniqueness
from src.validation.address import validate_address, validate_address_column
from src.validation.general import validate_general_columns, validate_and_correct_state
from src.validation.coordinates import validate_coordinates
from src.validation.smarty_validation import process_smarty_corrections

def validate_csv_column_count(input_csv, header_row_idx):
    """
    Validate that all rows in CSV have the same number of columns as the header.

    Returns:
        tuple: (is_valid, error_rows, all_rows_data)
        - is_valid: bool - True if all rows have correct column count
        - error_rows: list of dicts with {row_num, expected, actual, preview}
        - all_rows_data: list of all rows (for Excel generation if needed)
    """
    error_rows = []
    all_rows_data = []
    expected_column_count = None

    try:
        with open(input_csv, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)

            # Skip rows before header
            for i in range(header_row_idx):
                next(reader, None)

            # Read header to get expected column count
            header = next(reader, None)
            if header is None:
                return False, [{"row_num": 0, "expected": 0, "actual": 0, "preview": "Empty file", "data": []}], []

            expected_column_count = len(header)
            all_rows_data.append(header)  # Store header
            debug_print(f"CSV validation: Header has {expected_column_count} columns")

            # Check each data row
            row_num = header_row_idx + 2  # Account for header being row after skipped rows, +1 for 1-indexed
            for row in reader:
                all_rows_data.append(row)  # Store all rows for Excel generation
                actual_column_count = len(row)

                if actual_column_count != expected_column_count:
                    # Create preview (first 150 chars of row)
                    preview = ",".join(row)[:150] + ("..." if len(",".join(row)) > 150 else "")
                    error_rows.append({
                        "row_num": row_num,
                        "expected": expected_column_count,
                        "actual": actual_column_count,
                        "preview": preview,
                        "data": row
                    })
                    debug_print(f"Column count mismatch at row {row_num}: expected {expected_column_count}, got {actual_column_count}")

                row_num += 1

        is_valid = len(error_rows) == 0
        return is_valid, error_rows, all_rows_data

    except UnicodeDecodeError:
        # Try latin1 encoding
        try:
            with open(input_csv, 'r', encoding='latin1', newline='') as f:
                reader = csv.reader(f)

                for i in range(header_row_idx):
                    next(reader, None)

                header = next(reader, None)
                if header is None:
                    return False, [{"row_num": 0, "expected": 0, "actual": 0, "preview": "Empty file", "data": []}], []

                expected_column_count = len(header)
                all_rows_data.append(header)

                row_num = header_row_idx + 2
                for row in reader:
                    all_rows_data.append(row)
                    actual_column_count = len(row)

                    if actual_column_count != expected_column_count:
                        preview = ",".join(row)[:150] + ("..." if len(",".join(row)) > 150 else "")
                        error_rows.append({
                            "row_num": row_num,
                            "expected": expected_column_count,
                            "actual": actual_column_count,
                            "preview": preview,
                            "data": row
                        })

                    row_num += 1

                is_valid = len(error_rows) == 0
                return is_valid, error_rows, all_rows_data
        except Exception as e:
            debug_print(f"Failed to validate CSV with latin1 encoding: {str(e)}")
            return False, [{"row_num": 0, "expected": 0, "actual": 0, "preview": f"Error: {str(e)}", "data": []}], []

    except Exception as e:
        debug_print(f"Failed to validate CSV column count: {str(e)}")
        return False, [{"row_num": 0, "expected": 0, "actual": 0, "preview": f"Error: {str(e)}", "data": []}], []


def create_column_count_error_excel(all_rows_data, error_rows, output_path, expected_column_count):
    """
    Create Excel file with error rows highlighted in red.

    Args:
        all_rows_data: list of all rows from CSV (including header)
        error_rows: list of error row dicts with row_num
        output_path: where to save the Excel file
        expected_column_count: number of columns that should be in each row
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Column Count Errors"

        # Define red fill for error rows
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Create set of error row numbers for quick lookup (adjust for header)
        error_row_nums = {err["row_num"] for err in error_rows}

        # Write header
        if all_rows_data:
            header = all_rows_data[0]
            for col_idx, value in enumerate(header, start=1):
                ws.cell(row=1, column=col_idx, value=value)

        # Write all data rows
        excel_row = 2  # Start after header
        csv_row_num = 2  # Track actual CSV row number (1-indexed, accounting for header)

        for row_data in all_rows_data[1:]:  # Skip header in data
            # Pad or truncate row to match expected column count
            normalized_row = row_data[:expected_column_count] + [''] * max(0, expected_column_count - len(row_data))

            for col_idx, value in enumerate(normalized_row, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)

                # Highlight entire row if it's an error row
                if csv_row_num in error_row_nums:
                    cell.fill = red_fill

            excel_row += 1
            csv_row_num += 1

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        debug_print(f"Created column count error Excel file: {output_path}")
        return True

    except Exception as e:
        debug_print(f"Failed to create column count error Excel: {str(e)}")
        return False


def normalize_column_names(df, errors):
    """Normalize column names using common variations mapping."""
    from src.config.settings import COLUMN_NAME_MAPPING
    
    renamed_columns = []
    original_columns = list(df.columns)
    
    # Create case-insensitive mapping
    mapping_lower = {k.lower().strip(): v for k, v in COLUMN_NAME_MAPPING.items()}
    
    # Check for conflicts (both correct and variation exist)
    input_columns_lower = [col.lower().strip() for col in df.columns]
    conflicts_resolved = []
    
    # Find columns to rename
    columns_to_rename = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower in mapping_lower:
            target_column = mapping_lower[col_lower]
            # Check if target already exists (prefer existing correct name)
            if target_column not in [col for col in df.columns]:
                columns_to_rename[col] = target_column
            else:
                conflicts_resolved.append(f"Kept existing '{target_column}' column, ignored '{col}'")
    
    # Apply renames
    if columns_to_rename:
        df.rename(columns=columns_to_rename, inplace=True)
        for old_name, new_name in columns_to_rename.items():
            renamed_columns.append(f"'{old_name}' → '{new_name}'")
            debug_print(f"Renamed column '{old_name}' to '{new_name}'")
    
    # Log all changes
    if renamed_columns:
        debug_print(f"Column renames applied: {', '.join(renamed_columns)}")
    if conflicts_resolved:
        debug_print(f"Column conflicts resolved: {', '.join(conflicts_resolved)}")
    
    return df, renamed_columns, conflicts_resolved

def find_header_row(input_csv):
    """
    Find the actual header row in a CSV file by scanning the first few rows.
    Returns the row number where headers are found, or None if not found.
    """
    from src.config.settings import EXPECTED_COLUMNS
    import pandas as pd
    
    MAX_ROWS_TO_CHECK = 7
    
    debug_print(f"Scanning first {MAX_ROWS_TO_CHECK} rows to locate header row...")
    
    try:
        # Read first few rows without assuming header location
        sample_df = pd.read_csv(input_csv, nrows=MAX_ROWS_TO_CHECK, header=None, encoding='utf-8', keep_default_na=False)
        debug_print(f"Read {len(sample_df)} sample rows for header detection")
    except UnicodeDecodeError:
        try:
            sample_df = pd.read_csv(input_csv, nrows=MAX_ROWS_TO_CHECK, header=None, encoding='latin1', keep_default_na=False)
            debug_print(f"Read {len(sample_df)} sample rows with latin1 encoding")
        except Exception as e:
            debug_print(f"Failed to read sample rows: {str(e)}")
            return None
    except Exception as e:
        debug_print(f"Failed to read sample rows: {str(e)}")
        return None
    
    # Check each row to see if it could be the header
    for row_idx in range(len(sample_df)):
        potential_headers = sample_df.iloc[row_idx].astype(str).str.strip().tolist()
        debug_print(f"Checking row {row_idx}: {potential_headers[:5]}...")  # Show first 5 columns
        
        # Create a temporary DataFrame to test column normalization
        temp_df = pd.DataFrame([potential_headers]).T.reset_index()
        temp_df.columns = ['temp_col_' + str(i) for i in range(len(temp_df.columns))]
        temp_df = temp_df.T
        temp_df.columns = potential_headers
        
        # Apply column normalization (same logic as main process)
        try:
            normalized_df, renamed_columns, conflicts_resolved = normalize_column_names(temp_df, [])
            
            # Check if all expected columns are present after normalization
            normalized_columns = [col.lower().strip() for col in normalized_df.columns]
            expected_columns_lower = [col.lower() for col in EXPECTED_COLUMNS]
            
            missing_columns = [col for col in expected_columns_lower if col not in normalized_columns]
            
            if not missing_columns:
                # Found complete header row!
                debug_print(f"✅ Header row found at row {row_idx}")
                debug_print(f"Original headers: {potential_headers}")
                if renamed_columns:
                    debug_print(f"Column renames that will be applied: {renamed_columns}")
                if conflicts_resolved:
                    debug_print(f"Column conflicts resolved: {conflicts_resolved}")
                return row_idx
            else:
                debug_print(f"❌ Row {row_idx} missing columns: {missing_columns}")
                
        except Exception as e:
            debug_print(f"❌ Row {row_idx} failed normalization test: {str(e)}")
            continue
    
    # No valid header row found
    debug_print(f"❌ No valid header row found in first {MAX_ROWS_TO_CHECK} rows")
    return None

def save_csv(df, path, errors, header_comment="# the python version is 1.3.0.2\n"):
    """Save DataFrame to CSV with header comment."""
    try:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(header_comment)
            df.to_csv(f, index=False)
            f.flush()  # Ensure data is written to OS buffer
            os.fsync(f.fileno())  # Force write to disk
        if os.path.isfile(path):
            debug_print(f"Successfully saved and synced to disk: {path}")
            return path
        else:
            errors.append({
                "Row": "N/A",
                "Column": "N/A",
                "Error": f"Failed to save {path}. File does not exist.",
                "Value": "N/A"
            })
            debug_print(f"Error: File {path} was not created")
            return None
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Error saving {path}: {str(e)}",
            "Value": "N/A"
        })
        debug_print(f"Exception saving {path}: {traceback.format_exc()}")
        return None


def get_error_priority_and_fill(error_msg, col_name):
    """
    Centralized function to determine error priority and corresponding Excel fill color.
    
    Args:
        error_msg (str): The error message
        col_name (str): The column name where the error occurred
        
    Returns:
        tuple: (priority_level, fill_color)
            priority_level: 1=RED (highest), 2=PINK (medium), 3=YELLOW (lowest)
            fill_color: The corresponding PatternFill object
    """
    # RED (Priority 1) - Critical errors that must be fixed
    if (error_msg.startswith("Required field:") or
        error_msg == "Address lacks leading number followed by street name" or
        error_msg.startswith("Invalid technology:") or
        error_msg == "Invalid State Abbreviation" or
        error_msg == "Removal after Invalid response from Smarty" or
        error_msg == "Invalid format" or
        (col_name in ["download", "upload", "voip_lines_quantity", "business_customer"] and
         ("must be a number" in error_msg or "must be positive" in error_msg or
          "must be a non-negative integer" in error_msg or "must be 0 or 1" in error_msg))):
        return (1, RED_FILL)
    
    # PINK (Priority 2) - Requires manual review
    elif (error_msg.startswith("Corrected address is still invalid") or 
          error_msg == "Smarty Validation Failed - Returned for Review"):
        return (2, PINK_FILL)
    
    # YELLOW (Priority 3) - Warnings and minor issues
    else:
        return (3, YELLOW_FILL)


def save_excel(df, path, errors, corrected_cells, flagged_cells, rows_to_exclude=None):
    """Save DataFrame to Excel with cell coloring."""
    try:
        # Filter out excluded rows at the very beginning
        if rows_to_exclude:
            original_len = len(df)
            df = df[~df["OrigRowNum"].isin(rows_to_exclude)].reset_index(drop=True)
            debug_print(f"Filtered out {original_len - len(df)} rows from Excel output. Remaining: {len(df)} rows")
        
        # Keep DataFrame in original order - OrigRowNum should already be sequential
        sorted_df = df.copy().reset_index(drop=True)
        sorted_df["OrigRowNum"] = sorted_df["OrigRowNum"].astype(int)
        sorted_df["zip"] = sorted_df["zip"].astype("string")

        # Validate and convert data types
        for idx, row in sorted_df.iterrows():
            orig_row = row["OrigRowNum"]
            for col in ["zip", "download", "upload", "voip_lines_quantity", "business_customer"]:
                val = row[col]
                if pd.notna(val):
                    try:
                        if col == "zip":
                            # Check if this is a GPS-only row (all address fields empty)
                            row_address = row.get('address', '')
                            row_city = row.get('city', '')
                            row_state = row.get('state', '')
                            row_zip = row.get('zip', '')

                            address_empty = pd.isna(row_address) or str(row_address).strip() == ""
                            city_empty = pd.isna(row_city) or str(row_city).strip() == ""
                            state_empty = pd.isna(row_state) or str(row_state).strip() == ""
                            zip_empty = pd.isna(row_zip) or str(row_zip).strip() == ""

                            # Skip ZIP validation for GPS-only rows
                            if address_empty and city_empty and state_empty and zip_empty:
                                debug_print(f"save_excel: Skipping ZIP validation for OrigRowNum={orig_row}: All address fields empty (GPS-only row)")
                                continue

                            if not re.match(r"^\d{5}(-\d{4})?$", str(val)):
                                sorted_df.loc[idx, col] = pd.NA
                                errors.append({
                                    "Row": orig_row,
                                    "Column": col,
                                    "Error": "Invalid ZIP code",
                                    "Value": val
                                })
                                corrected_cells[(idx, col)] = {
                                    "row": int(sorted_df["OrigRowNum"].iloc[idx]),
                                    "original": val,
                                    "corrected": None,
                                    "type": "Invalid ZIP Replacement",
                                    "status": "Valid"
                                }
                        elif col in ["download", "upload"]:
                            float_val = float(val)
                            sorted_df.loc[idx, col] = float_val
                            if str(val) != str(float_val):
                                corrected_cells[(idx, col)] = {
                                    "row": int(sorted_df["OrigRowNum"].iloc[idx]),
                                    "original": val,
                                    "corrected": float_val,
                                    "type": f"{col.capitalize()} Format Conversion",
                                    "status": "Valid"
                                }
                        elif col == "voip_lines_quantity":
                            int_val = int(float(val))
                            sorted_df.loc[idx, col] = int_val
                            if str(val) != str(int_val):
                                corrected_cells[(idx, col)] = {
                                    "row": int(sorted_df["OrigRowNum"].iloc[idx]),
                                    "original": val,
                                    "corrected": int_val,
                                    "type": "VoIP Lines Format Conversion",
                                    "status": "Valid"
                                }
                        elif col == "business_customer":
                            int_val = int(val)
                            if int_val not in [0, 1]:
                                sorted_df.loc[idx, col] = pd.NA
                                errors.append({
                                    "Row": orig_row,
                                    "Column": col,
                                    "Error": "Business customer must be 0 or 1",
                                    "Value": val
                                })
                                corrected_cells[(idx, col)] = {
                                    "row": int(sorted_df["OrigRowNum"].iloc[idx]),
                                    "original": val,
                                    "corrected": None,
                                    "type": "Invalid Business Customer Replacement",
                                    "status": "Valid"
                                }
                            else:
                                sorted_df.loc[idx, col] = int_val
                    except (ValueError, TypeError) as e:
                        sorted_df.loc[idx, col] = pd.NA
                        errors.append({
                            "Row": orig_row,
                            "Column": col,
                            "Error": f"Invalid {col} value: {str(e)}",
                            "Value": val
                        })
                        corrected_cells[(idx, col)] = {
                            "row": int(sorted_df["OrigRowNum"].iloc[idx]),
                            "original": val,
                            "corrected": None,
                            "type": f"Invalid {col.capitalize()} Replacement",
                            "status": "Valid"
                        }

        # Create mapping from OrigRowNum to Excel row number
        orig_row_to_excel_row = {row["OrigRowNum"]: i + 2 for i, row in sorted_df.iterrows()}
        
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sorted_df.to_excel(writer, sheet_name="Corrected Data", index=False)
            wb = writer.book
            ws = wb["Corrected Data"]
            col_map = {col: idx + 1 for idx, col in enumerate(sorted_df.columns)}

            for idx, row in sorted_df.iterrows():
                excel_row = idx + 2
                for col, val in row.items():
                    excel_col = openpyxl.utils.get_column_letter(col_map[col])
                    cell = ws[f"{excel_col}{excel_row}"]
                    cell.value = "" if pd.isna(val) else val
                    if col == "voip_lines_quantity":
                        cell.number_format = "0"

            # Track which cells have been filled and their priority levels
            cell_fills = {}  # {(excel_row, excel_col): (priority_level, fill_color)}
            
            # Process flagged errors and determine priorities
            for (row_idx, col_name), cell_data in flagged_cells.items():
                # Handle both old format (just error_msg) and new format (error_msg, orig_row)
                if isinstance(cell_data, tuple):
                    error_msg, orig_row_stored = cell_data
                else:
                    # Fallback for old format
                    error_msg = cell_data
                    orig_row_stored = None
                    if row_idx < len(sorted_df):
                        orig_row_stored = sorted_df["OrigRowNum"].iloc[row_idx]
                
                # Only process if this row still exists in filtered DataFrame
                if orig_row_stored and orig_row_stored in orig_row_to_excel_row:
                    excel_row = orig_row_to_excel_row[orig_row_stored]
                    
                    if col_name in col_map:
                        excel_col = openpyxl.utils.get_column_letter(col_map[col_name])
                        cell_key = (excel_row, excel_col)
                        
                        # Get priority and fill color using centralized function
                        priority_level, fill_color = get_error_priority_and_fill(error_msg, col_name)
                        
                        # Apply fill only if this is higher priority (lower number) than existing
                        if cell_key not in cell_fills or priority_level < cell_fills[cell_key][0]:
                            cell_fills[cell_key] = (priority_level, fill_color)
                            debug_print(f"Set priority {priority_level} fill for error: '{error_msg}' at OrigRowNum {orig_row_stored}, Excel row {excel_row}, col {excel_col}")
                        else:
                            debug_print(f"Skipped lower priority fill for error: '{error_msg}' at OrigRowNum {orig_row_stored}, Excel row {excel_row}, col {excel_col}")
                    else:
                        debug_print(f"Column {col_name} not found in col_map for orig_row {orig_row_stored}")
                else:
                    debug_print(f"Skipping fill for orig_row={orig_row_stored}, error_msg={error_msg}: Row was excluded from Excel output")

            # Apply all determined fills
            for (excel_row, excel_col), (priority_level, fill_color) in cell_fills.items():
                ws[f"{excel_col}{excel_row}"].fill = fill_color
                priority_name = {1: "RED", 2: "PINK", 3: "YELLOW"}[priority_level]
                debug_print(f"Applied {priority_name} fill at Excel row {excel_row}, col {excel_col}")

            # Apply correction fills using OrigRowNum for robust lookup
            for (row_idx, col_name), info in corrected_cells.items():
                if info["status"] == "Valid":
                    # Use the OrigRowNum stored in the correction info for accurate mapping
                    orig_row = info.get("row")
                    
                    # Only process if this OrigRowNum still exists in the filtered DataFrame
                    if orig_row and orig_row in orig_row_to_excel_row:
                        excel_row = orig_row_to_excel_row[orig_row]
                        if col_name in col_map:
                            excel_col = openpyxl.utils.get_column_letter(col_map[col_name])
                            cell_key = (excel_row, excel_col)
                            # Only apply GREEN fill if no error fill has been applied
                            if cell_key not in cell_fills:
                                ws[f"{excel_col}{excel_row}"].fill = GREEN_FILL
                                debug_print(f"Applied GREEN correction fill at Excel row {excel_row}, col {excel_col}")
                            else:
                                debug_print(f"Skipped GREEN fill (error fill takes precedence) at Excel row {excel_row}, col {excel_col}")
                        else:
                            debug_print(f"Failed to apply correction fill for orig_row={orig_row}: Could not map to Excel row")
                    else:
                        debug_print(f"Skipping correction fill for orig_row={orig_row}: Row was excluded from Excel output")

        if os.path.isfile(path):
            debug_print(f"Successfully saved: {path}")
            return path
        else:
            errors.append({
                "Row": "N/A",
                "Column": "",
                "Error": f"Failed to save {path}. File does not exist.",
                "Value": ""
            })
            debug_print(f"Error: File {path} was not created")
            return None
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "Error",
            "Error": f"Error saving {path}: {str(e)}",
            "Value": ""
        })
        debug_print(f"Exception saving {path}: {str(traceback.format_exc())}")
        return None

def validate_subscriber_file(input_csv, company_id, period):
    """Validate subscriber CSV and generate output files."""
    from src.utils.reporting import generate_validation_report
    import time
    import shutil
    errors = []
    corrected_cells = {}
    flagged_cells = {}
    pobox_errors = []
    rows_to_remove = []
    non_unique_row_removals = []
    start_time = time.time()

    # Build new directory structure: /var/www/broadband/Subscriber_File_Validations/{period}/{company_id}/
    base_validation_dir = "/var/www/broadband/Subscriber_File_Validations"
    output_dir = os.path.join(base_validation_dir, period, company_id)

    # Delete existing directory if it exists (replacing old validation for same period)
    if os.path.exists(output_dir):
        debug_print(f"Removing existing validation directory: {output_dir}")
        shutil.rmtree(output_dir)

    # Create fresh directory structure
    os.makedirs(output_dir, exist_ok=True)
    debug_print(f"Created output directory: {output_dir}")

    # Use output_dir instead of company_id for all file operations
    company_id = output_dir  # Override company_id to use new path

    # Create company directory
    try:
        if os.path.exists(company_id):
            shutil.rmtree(company_id)
        os.makedirs(company_id)
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Failed to create directory {company_id}: {str(e)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, os.path.basename(input_csv))
        return

    # Copy input CSV
    original_filename = os.path.basename(input_csv)
    base_filename = os.path.splitext(original_filename)[0]
    output_original_csv = os.path.join(company_id, original_filename)
    try:
        shutil.copyfile(input_csv, output_original_csv)
        debug_print(f"Copied input CSV to {output_original_csv}")
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Failed to copy {input_csv}: {str(e)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Find the actual header row (handles files with extra rows at top)
    header_row_idx = find_header_row(input_csv)
    
    if header_row_idx is None:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Could not locate valid column headers in first 7 rows of {input_csv}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # NEW: Validate column count consistency and auto-fix by truncating extra columns
    debug_print(f"Validating CSV column count consistency...")
    is_valid, error_rows, all_rows_data = validate_csv_column_count(input_csv, header_row_idx)

    # If column count issues exist, create a cleaned CSV with truncated rows
    csv_to_process = input_csv  # Default to original

    if not is_valid:
        debug_print(f"Column count validation found {len(error_rows)} rows with mismatched column counts - truncating to {len(all_rows_data[0])} columns")

        # Create a cleaned CSV with all rows truncated to correct column count
        cleaned_csv_path = os.path.join(company_id, f"{base_filename}_cleaned_temp.csv")
        expected_column_count = len(all_rows_data[0])

        try:
            with open(cleaned_csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in all_rows_data:
                    # Truncate to expected column count (removes extra columns)
                    # Pad with empty strings if row is too short
                    normalized_row = row[:expected_column_count] + [''] * max(0, expected_column_count - len(row))
                    writer.writerow(normalized_row)

            debug_print(f"Created cleaned CSV with truncated rows: {cleaned_csv_path}")
            csv_to_process = cleaned_csv_path  # Process the cleaned version

            # Log which rows were affected (for debugging)
            debug_print(f"Truncated rows: {[err['row_num'] for err in error_rows]}")

        except Exception as e:
            debug_print(f"Failed to create cleaned CSV: {str(e)}")
            errors.append({
                "Row": "N/A",
                "Column": "N/A",
                "Error": f"Failed to process CSV with column count mismatches: {str(e)}",
                "Value": "N/A"
            })
            save_errors_and_exit(errors, company_id, original_filename)
            return
    else:
        debug_print(f"Column count validation passed - all rows have consistent column counts")

    # Skip rows above the header and read CSV properly
    # Note: If we created a cleaned CSV, header_row_idx should be 0 (cleaned CSV has no extra rows)
    skiprows = list(range(header_row_idx)) if header_row_idx > 0 and csv_to_process == input_csv else None

    # Read input CSV starting from the correct header row (or cleaned CSV if created)
    try:
        df = pd.read_csv(csv_to_process, dtype={col: str for col in EXPECTED_COLUMNS}, encoding='utf-8', keep_default_na=False, sep=',', skiprows=skiprows)
        debug_print(f"Read CSV successfully: {len(df)} rows (skipped {header_row_idx if skiprows else 0} header rows)")
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(csv_to_process, dtype={col: str for col in EXPECTED_COLUMNS}, encoding='latin1', keep_default_na=False, sep=',', skiprows=skiprows)
            debug_print(f"Read CSV with latin1 encoding: {len(df)} rows (skipped {header_row_idx if skiprows else 0} header rows)")
        except Exception as e:
            errors.append({
                "Row": "N/A",
                "Column": "N/A",
                "Error": f"Failed to read CSV {csv_to_process}: {str(e)}",
                "Value": "N/A"
            })
            save_errors_and_exit(errors, company_id, original_filename)
            return
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Failed to read CSV {csv_to_process}: {str(e)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Normalize column names before validation
    df, renamed_columns, conflicts_resolved = normalize_column_names(df, errors)
    
    # Log column name changes for reporting
    if renamed_columns or conflicts_resolved:
        column_changes_msg = []
        if renamed_columns:
            column_changes_msg.append(f"Renamed: {', '.join(renamed_columns)}")
        if conflicts_resolved:
            column_changes_msg.append(f"Conflicts: {', '.join(conflicts_resolved)}")
        debug_print(f"Column normalization completed: {'; '.join(column_changes_msg)}")

    # Insert OrigRowNum - account for skipped header rows
    df.insert(0, "OrigRowNum", range(header_row_idx + 2, header_row_idx + 2 + len(df)))

    # Validate columns
    input_columns = df.columns.str.lower().tolist()
    missing_columns = [col for col in EXPECTED_COLUMNS if col.lower() not in input_columns]
    if missing_columns:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Missing columns: {', '.join(missing_columns)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Create cleaned DataFrame
    column_mapping = {col: col.lower() for col in df.columns if col.lower() in EXPECTED_COLUMNS}
    column_mapping['OrigRowNum'] = 'OrigRowNum'
    output_columns = ["OrigRowNum"] + EXPECTED_COLUMNS
    cleaned_df = df[list(column_mapping.keys())].rename(columns=column_mapping)[output_columns]

    # Early data cleaning: Remove commas from customer column
    if 'customer' in cleaned_df.columns:
        original_customer_values = cleaned_df['customer'].copy()
        cleaned_df['customer'] = cleaned_df['customer'].astype(str).str.replace(',', '', regex=False)
    
        # Track corrections for customers where commas were removed
        for idx, (original, cleaned) in enumerate(zip(original_customer_values, cleaned_df['customer'])):
            if pd.notna(original) and str(original) != str(cleaned) and ',' in str(original):
                corrected_cells[(idx, "customer")] = {
                    "row": int(cleaned_df["OrigRowNum"].iloc[idx]),
                    "original": str(original),
                    "corrected": str(cleaned),
                    "type": "Comma Removal",
                    "status": "Valid"
                }
                debug_print(f"Removed comma from customer ID for OrigRowNum={cleaned_df['OrigRowNum'].iloc[idx]}: '{original}' -> '{cleaned}'")

    # Early data cleaning: Remove commas from address column
    if 'address' in cleaned_df.columns:
        original_address_values = cleaned_df['address'].copy()
        cleaned_df['address'] = cleaned_df['address'].astype(str).str.replace(',', '', regex=False)
    
        # Track corrections for addresses where commas were removed
        for idx, (original, cleaned) in enumerate(zip(original_address_values, cleaned_df['address'])):
            if pd.notna(original) and str(original) != str(cleaned) and ',' in str(original):
                corrected_cells[(idx, "address")] = {
                    "row": int(cleaned_df["OrigRowNum"].iloc[idx]),
                    "original": str(original),
                    "corrected": str(cleaned),
                    "type": "Early Comma Removal",
                    "status": "Valid"
                }
                debug_print(f"Removed comma from address for OrigRowNum={cleaned_df['OrigRowNum'].iloc[idx]}: '{original}' -> '{cleaned}'")

    # *** NEW SECTION: Early data cleaning: Remove commas from numeric columns ***
    numeric_columns = ['download', 'upload', 'voip_lines_quantity', 'zip']
    for col in numeric_columns:
        if col in cleaned_df.columns:
            original_values = cleaned_df[col].copy()
            cleaned_df[col] = cleaned_df[col].astype(str).str.replace(',', '', regex=False)
        
            # Track corrections for values where commas were removed
            for idx, (original, cleaned) in enumerate(zip(original_values, cleaned_df[col])):
                if pd.notna(original) and str(original) != str(cleaned) and ',' in str(original):
                    corrected_cells[(idx, col)] = {
                        "row": int(cleaned_df["OrigRowNum"].iloc[idx]),
                        "original": str(original),
                        "corrected": str(cleaned),
                        "type": f"{col.capitalize()} Comma Removal",
                        "status": "Valid"
                    }
                    debug_print(f"Removed comma from {col} for OrigRowNum={cleaned_df['OrigRowNum'].iloc[idx]}: '{original}' -> '{cleaned}'")

    # Early data cleaning: Convert technology column to lowercase
    if 'technology' in cleaned_df.columns:
        original_technology_values = cleaned_df['technology'].copy()
        cleaned_df['technology'] = cleaned_df['technology'].astype(str).str.lower().str.strip()
    
        # Track corrections for technology values that were changed
        for idx, (original, cleaned) in enumerate(zip(original_technology_values, cleaned_df['technology'])):
            if pd.notna(original) and str(original) != str(cleaned):
                corrected_cells[(idx, "technology")] = {
                    "row": int(cleaned_df["OrigRowNum"].iloc[idx]),
                    "original": str(original),
                    "corrected": str(cleaned),
                    "type": "Early Technology Case Normalization",
                    "status": "Valid"
                }
                debug_print(f"Converted technology to lowercase for OrigRowNum={cleaned_df['OrigRowNum'].iloc[idx]}: '{original}' -> '{cleaned}'")

    # Apply data type conversions
    for col, dtype in DTYPE_DICT.items():
        if col in cleaned_df.columns:
            try:
                cleaned_df[col] = cleaned_df[col].astype(dtype)
            except Exception as e:
                debug_print(f"Failed to cast {col} to {dtype}: {str(e)}")
    cleaned_df["address"] = cleaned_df["address"].fillna("").astype(str).str.upper()
    cleaned_df["city"] = cleaned_df["city"].fillna("").astype(str).str.upper()

    # Phase 1: Non-standard address endings and state correction
    for idx, val in enumerate(cleaned_df["address"].fillna("").astype(str).str.strip()):
        orig_row = cleaned_df["OrigRowNum"][idx]

        # Check if ALL address fields are empty (GPS-only row)
        address_empty = not val or val.strip() == ""
        city_empty = pd.isna(cleaned_df["city"].iloc[idx]) or str(cleaned_df["city"].iloc[idx]).strip() == ""
        state_empty = pd.isna(cleaned_df["state"].iloc[idx]) or str(cleaned_df["state"].iloc[idx]).strip() == ""
        zip_empty = pd.isna(cleaned_df["zip"].iloc[idx]) or str(cleaned_df["zip"].iloc[idx]).strip() == ""

        # If ALL address fields are empty, skip address validation (GPS-only row)
        if address_empty and city_empty and state_empty and zip_empty:
            debug_print(f"Phase 1: Skipping address validation for OrigRowNum={orig_row}: All address fields empty (GPS-only row)")
            continue

        state = cleaned_df["state"][idx]  # NEW: Get state from the row
        validate_address(val, orig_row, idx, errors, corrected_cells, flagged_cells, pobox_errors, rows_to_remove, non_standard_only=True, state=state)
        if (idx, "address") in corrected_cells and corrected_cells[(idx, "address")]["status"] == "Valid":
            cleaned_df.loc[idx, "address"] = corrected_cells[(idx, "address")]["corrected"]

    # State validation
    for idx, (state_val, zip_val) in enumerate(zip(cleaned_df["state"].fillna(""), cleaned_df["zip"].fillna(""))):
        orig_row = cleaned_df["OrigRowNum"][idx]

        # Check if ALL address fields are empty (GPS-only row) - skip state validation
        address_empty = pd.isna(cleaned_df["address"].iloc[idx]) or str(cleaned_df["address"].iloc[idx]).strip() == ""
        city_empty = pd.isna(cleaned_df["city"].iloc[idx]) or str(cleaned_df["city"].iloc[idx]).strip() == ""
        state_empty = pd.isna(cleaned_df["state"].iloc[idx]) or str(cleaned_df["state"].iloc[idx]).strip() == ""
        zip_empty = pd.isna(cleaned_df["zip"].iloc[idx]) or str(cleaned_df["zip"].iloc[idx]).strip() == ""

        if address_empty and city_empty and state_empty and zip_empty:
            debug_print(f"Skipping state validation for OrigRowNum={orig_row}: All address fields empty (GPS-only row)")
            continue

        corrected_state = validate_and_correct_state(state_val, zip_val, idx, orig_row, errors, corrected_cells, flagged_cells)
        if (idx, "state") in corrected_cells and corrected_cells[(idx, "state")]["status"] == "Valid":
            cleaned_df.loc[idx, "state"] = corrected_cells[(idx, "state")]["corrected"]

    # Ensure cleaned_df is in the correct state for Phase 2
    cleaned_df["OrigRowNum"] = cleaned_df["OrigRowNum"].astype(int)
    cleaned_df.reset_index(drop=True, inplace=True)

    # Phase 2: Comprehensive validation (keeping all rows in DataFrame)
    validate_general_columns(cleaned_df, errors, corrected_cells, flagged_cells)
    validate_coordinates(cleaned_df, errors, corrected_cells, flagged_cells)
    validate_address_column(cleaned_df, errors, corrected_cells, flagged_cells, pobox_errors, rows_to_remove)

    # Keep DataFrame in original input order - no sorting needed
    # This eliminates index confusion and preserves input file order
    debug_print("Maintaining original input order - no sorting applied")

    # Validate customer uniqueness (marks duplicates for removal but doesn't remove them)
    # NOTE: Only handles duplicate customer IDs, not data-based duplicates
    # Different customers with same data are KEPT (common in MDUs - same building, different apartments)
    validate_customer_uniqueness(cleaned_df, errors, rows_to_remove, non_unique_row_removals, corrected_cells, flagged_cells)
    save_csv(pd.DataFrame(non_unique_row_removals), os.path.join(company_id, f"{base_filename}_N_Unq_Row_Remvl.csv"), errors)

    # Deduplicate errors before further processing
    unique_errors = {f"{e['Row']}_{e['Column']}_{e['Error']}_{e['Value']}": e for e in errors}
    errors = list(unique_errors.values())

    # Convert address patterns
    for idx, address in enumerate(cleaned_df["address"].fillna("").astype(str)):
        orig_row = cleaned_df["OrigRowNum"][idx]
        corrected_address = re.sub(r"(?i)\bFarm\s+to\s+Market\b", "FM", address)
        corrected_address = re.sub(r"(?i)\bCounty\s+(?:Road|Rd)\b", "CR", corrected_address)
        corrected_address = re.sub(r"(?i)\bPrivate\s+Road\b", "PVT RD", corrected_address)
        if corrected_address != address:
            cleaned_df.loc[idx, "address"] = corrected_address
            corrected_cells[(idx, "address")] = {
                "row": int(cleaned_df["OrigRowNum"].iloc[idx]),
                "original": address,
                "corrected": corrected_address,
                "type": "Address Pattern Conversion",
                "status": "Valid"
            }

    # NEW LOGIC: Always run Smarty if there are Smarty-eligible addresses
    # Only skip Smarty entirely if there are address field critical errors (missing address, city, or state)
    # that would prevent address validation. Other critical errors (missing download, upload, customer, etc.)
    # should not prevent us from fixing addresses that CAN be validated.

    # Count address-field critical errors only (address, city, state - not zip since Smarty can fill that in)
    address_critical_errors = [error for error in errors if
                               error.get("Error", "").startswith("Required field:") and
                               error.get("Column", "") in ["address", "city", "state"]]
    address_critical_flagged = [(k, v) for k, v in flagged_cells.items() if
                                ((isinstance(v, tuple) and v[0].startswith("Required field:")) or
                                 (isinstance(v, str) and v.startswith("Required field:"))) and
                                k[1] in ["address", "city", "state"]]

    # Check if there are ANY Smarty-eligible addresses in flagged_cells
    from src.validation.smarty_validation import should_send_to_smarty
    smarty_eligible_count = sum(1 for (row_idx, col_name), cell_data in flagged_cells.items()
                                if should_send_to_smarty(cell_data[0] if isinstance(cell_data, tuple) else cell_data))

    if smarty_eligible_count == 0:
        debug_print("No addresses eligible for Smarty processing - skipping Smarty API")
        smarty_results = {
            'addresses_sent': 0,
            'successful_corrections': 0,
            'failed_corrections': 0,
            'action_taken': 'NO_ELIGIBLE_ADDRESSES',
            'smarty_corrections': [],
            'processing_time': 0.0
        }
    elif address_critical_errors or address_critical_flagged:
        # Only log that some rows have address field issues, but still process what we can
        debug_print(f"Found {len(address_critical_errors)} address field critical errors, but proceeding with Smarty for {smarty_eligible_count} eligible addresses")
        debug_print("Starting Smarty API processing for eligible addresses...")
        smarty_results = process_smarty_corrections(
            cleaned_df, errors, corrected_cells, flagged_cells, company_id, base_filename
        )
    else:
        # No address field critical errors - proceed normally
        debug_print(f"Starting Smarty API processing for {smarty_eligible_count} eligible addresses...")
        smarty_results = process_smarty_corrections(
            cleaned_df, errors, corrected_cells, flagged_cells, company_id, base_filename
        )

    # Collect all rows to exclude from final Excel output (immediate removals only)
    all_rows_to_exclude = set(rows_to_remove).union(
        int(error["Row"]) for error in pobox_errors
    )
    # Note: No Smarty-based removals added here - final decision made in reporting

    debug_print(f"Rows to remove from duplicates: {rows_to_remove}")
    debug_print(f"Total rows to exclude from Excel: {len(all_rows_to_exclude)} (PO Box: {len(pobox_errors)}, Invalid/Duplicates: {len(rows_to_remove)})")
    debug_print("Smarty failures will be evaluated by final validation decision logic")

    # Save final outputs (rows are filtered only in save_excel)
    save_excel(cleaned_df, os.path.join(company_id, f"{base_filename}_Corrected_Subscribers.xlsx"), errors, corrected_cells, flagged_cells, all_rows_to_exclude)
    save_csv(pd.DataFrame(errors), os.path.join(company_id, f"{base_filename}_Errors.csv"), errors)
    save_csv(pd.DataFrame(pobox_errors), os.path.join(company_id, f"{base_filename}_POBox_Errors.csv"), errors)

    # Save clean CSV for Code B geocoding (matching Excel content - excluded rows removed, OrigRowNum stripped)
    # This CSV is ready for Phase 2 processing (geocoding, tract assignment, database insertion)
    cleaned_for_csv = cleaned_df[~cleaned_df["OrigRowNum"].isin(all_rows_to_exclude)].copy()
    # Remove OrigRowNum column - not needed for geocoding phase
    if 'OrigRowNum' in cleaned_for_csv.columns:
        cleaned_for_csv = cleaned_for_csv.drop(columns=['OrigRowNum'])
    # IMPORTANT: Use empty header_comment for Code B compatibility (no python version comment)
    save_csv(cleaned_for_csv, os.path.join(company_id, f"{base_filename}_Corrected_Subscribers.csv"), errors, header_comment="")
    debug_print(f"Saved clean CSV for geocoding: {len(cleaned_for_csv)} rows (excluded {len(all_rows_to_exclude)} problematic rows)")
    
    # Generate validation report and get file validation status
    # This is where the final decision on Smarty failures is made using subscriber-count thresholds
    vr_excel_path, vr_json_path, file_validation = generate_validation_report(
        cleaned_df, company_id, base_filename, errors, start_time, corrected_cells, 
        flagged_cells, pobox_errors, non_unique_row_removals, smarty_results
    )
    
    # Log final validation status
    debug_print(f"=== FINAL VALIDATION STATUS ===")
    debug_print(f"File Status: {file_validation.get('file_status', 'Unknown')}")
    debug_print(f"Validation Reason: {file_validation.get('validation_reason', 'No reason provided')}")
    debug_print(f"Total Subscribers: {file_validation.get('total_subscribers', 0)}")
    debug_print(f"Address Errors: {file_validation.get('address_error_count', 0)} ({file_validation.get('address_error_percentage', 0):.2f}%)")
    debug_print(f"Non-Address Errors: {file_validation.get('non_address_error_count', 0)}")
    debug_print(f"Requires Manual Review: {file_validation.get('requires_manual_review', True)}")
    debug_print(f"Threshold Used: {file_validation.get('threshold_used', {}).get('description', 'Unknown')}")
    debug_print(f"================================")
    
    # Display user-friendly status message
    if file_validation.get('file_status') == 'Valid':
        print(f"\n✅ SUCCESS: File validation completed successfully!")
        print(f"📄 Status: {file_validation['file_status']}")
        print(f"📊 Final subscriber count: {file_validation.get('total_subscribers', 0)}")
        if file_validation.get('address_error_count', 0) > 0:
            print(f"🔧 Automatically removed {file_validation['address_error_count']} rows with address issues")
        print(f"🎯 File is ready for FCC BDC submission")
        print(f"📂 Output: {company_id}/{base_filename}_Corrected_Subscribers.xlsx")
    else:
        print(f"\n❌ ATTENTION: File requires manual review")
        print(f"📄 Status: {file_validation['file_status']}")
        print(f"⚠️  Reason: {file_validation.get('validation_reason', 'Unknown issue')}")
        print(f"📊 Total subscribers: {file_validation.get('total_subscribers', 0)}")
        if file_validation.get('non_address_error_count', 0) > 0:
            print(f"🚨 Critical errors in required fields: {file_validation['non_address_error_count']} rows")
        if file_validation.get('address_error_count', 0) > 0:
            print(f"📍 Address issues requiring review: {file_validation['address_error_count']} rows ({file_validation.get('address_error_percentage', 0):.2f}%)")
        print(f"👤 Manual intervention required before FCC BDC submission")
        print(f"📂 Review file: {company_id}/{base_filename}_Corrected_Subscribers.xlsx")
        print(f"📋 Detailed report: {company_id}/{base_filename}_VR.xlsx")
    
    print(f"\n📊 Validation Report: {company_id}/{base_filename}_VR.xlsx")
    print(f"📄 Processing completed in {time.time() - start_time:.2f} seconds")

    return cleaned_df, errors, file_validation

def save_errors_and_exit(errors, company_id, base_filename):
    """Save errors and exit the program."""
    import sys
    path = os.path.join(company_id, f"{base_filename}_Errors.csv")
    save_csv(pd.DataFrame(errors), path, errors)
    debug_print(f"Errors saved to {path}. Exiting.")
    sys.exit(1)