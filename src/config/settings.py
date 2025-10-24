"""Configuration settings and constants for subscriber validation."""

import os
import re
from openpyxl.styles import PatternFill

SMARTY_BATCH_SIZE = 100  # Maximum addresses per batch (Smarty limit)
SMARTY_MIN_BATCH_SIZE = 5  # Use batching only if we have this many addresses
SMARTY_BATCH_TIMEOUT = 60  # Longer timeout for batch requests
SMARTY_BATCH_MAX_PAYLOAD_BYTES = 32768  # Smarty's max payload size (32KB) for batch requests
SMARTY_MAX_RETRIES = 3  # Number of retry attempts for API calls
SMARTY_RATE_LIMIT_DELAY = 0.1  # Seconds between requests for rate limiting
SMARTY_TIMEOUT_SECONDS = 30  # Timeout for single-address requests

# Environment variables
SMARTY_AUTH_ID = os.getenv("SMARTY_AUTH_ID")
SMARTY_AUTH_TOKEN = os.getenv("SMARTY_AUTH_TOKEN")

# Constants
DEBUG_MODE = True
VALID_STATES = [
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
    "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
    "VA", "WA", "WV", "WI", "WY", "DC", "PR", "VI", "GU", "AS", "MP"
]
VALID_TECHNOLOGIES = [
    "wireless_unlicensed", "wireless_gaa", "wireless_pal", "wireless_educational",
    "fiber", "cable", "adsl2", "ethernet", "voip"
]
EXPECTED_COLUMNS = [
    "customer", "lat", "lon", "address", "city", "state", "zip", "download", "upload",
    "voip_lines_quantity", "business_customer", "technology"
]
SMARTY_USAGE_LOG_PATH = "./smarty_logs/Smarty_Usage_Log.csv"

# Street ending patterns
MULTI_WORD_ENDINGS = (
    r"\bUS Highway\b|\bUS Hwy\b|\bPrivate Road\b|\bCounty Road\b|\bCounty Rd\b|\bCo Rd\b|\bState Route\b|"
    r"\bFarm to Market\b|\bCounty Hwy \d+\b|\bCounty FM \d+\b|\bFM Road \d+\b|"
    r"\bFire District \d+ Rd\b|\bState Hwy \d+\b|\bKamehameha Hwy\b|\bMamalahoa Hwy\b|"
    r"\bRoute C-\d+\b|\bRoute [A-Z]{2}\b|\b[A-Z]{2} Road\b|\bRS \d+\b|\bKY RS \d+\b"
    r"|\bLoop \d+(?: (?:N|S|E|W|NE|NW|SE|SW))?\b"
)
SINGLE_WORD_ENDINGS = (
    r"\sAlley\b|\sALY\b|\sAvenue\b|\sAve\b|\sAv\b|\sBoulevard\b|\sBlvd\b|\sCircle\b|\sCir\b|\sCr\b|"
    r"\sCourt\b|\sCt\b|\sDrive\b|\sDr\b|\sExpressway\b|\sExpy\b|\sLane\b|\sLn\b|\sLoop\b|"
    r"\sParkway\b|\sPkwy\b|\sPlace\b|\sPl\b|\sRoad\b|\sRd\b|\sSquare\b|\sSq\b|\sStreet\b|\sSt\b|"
    r"\sTerrace\b|\sTer\b|\sTrail\b|\sTrl\b|\sWay\b|\sWy\b|\sShores\b|\sCreek\b|\sCrk\b|"
    r"|\sLoop \d+(?: (?:N|S|E|W|NE|NW|SE|SW))?\b"
)
SPECIFIC_ROAD_PATTERN = (
    r"(?i)(?:\d+\s+)?(?:County\s*(?:Road|Rd|CR)|Private\s*Road|Us\s*Hwy|Hwy\s*\d+|"
    r"HWY\s*\d+|Highway\s*\d+|Farm\s*to\s*Market|Farm\s*Road|Farm\s*to\s*Market\s*Road|"
    r"FM\s*Rd|FM\s+\d+|State\s*(?:Road|Rd|Route)|Old\s*State\s*(?:Road|Rd)|"
    r"\b(?:FM|CR|SH|TX|HWY|US)-\d+\b|"
    r"(?:AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|DC|PR|VI|GU|AS|MP)-\d+|"
    r"(?:Alabama|Alaska|Arizona|Arkansas|California|Colorado|Connecticut|Delaware|Florida|Georgia|Hawaii|Idaho|Illinois|Indiana|Iowa|Kansas|Kentucky|Louisiana|Maine|Maryland|Massachusetts|Michigan|Minnesota|Mississippi|Missouri|Montana|Nebraska|Nevada|New\sHampshire|New\sJersey|New\sMexico|New\sYork|North\sCarolina|North\sDakota|Ohio|Oklahoma|Oregon|Pennsylvania|Rhode\sIsland|South\sCarolina|South\sDakota|Tennessee|Texas|Utah|Vermont|Virginia|Washington|West\sVirginia|Wisconsin|Wyoming|District\sof\sColumbia|Puerto\sRico|Virgin\sIslands|Guam|American\sSamoa|Northern\sMariana\sIslands)\s*(?:Hwy|Highway|Route|Rte|Rt)\s*\d+)"
    r"\s*(?:\d+\s*(?:North|South|East|West|Northeast|Northwest|Southeast|Southwest|N|S|E|W|NE|NW|SE|SW)\b)?\b"
)
STREET_ENDINGS = f"({MULTI_WORD_ENDINGS})|({SINGLE_WORD_ENDINGS})"
PO_BOX = r"\b(?:PO Box|P\.O\. Box|Post Office Box|P\s*O\s*Box|POBox|P\.O\.Box)\b"
RURAL_ROUTES = r"\bRR \d+ Box \d+\b|\bRural Route \d+ Box \d+\b|\bR\.R\. \d+ Box \d+\b|\bHC \d+ Box \d+\b"
FORBIDDEN_CHARS = r'[!@#$%^&*()+={}[\]|\"?:;<,>]'
NON_STANDARD_ENDINGS = r"(?i)(?:^|\s)(Apt|Apartment|Suite|Ste|Unit|Room|Rm|Floor|Fl|Building|Bldg|Dept|Ofc|Lot|Slip|Space|Hangar|Box)(?:\s*[A-Z0-9][\w\-]*)?(?=\s*$)|(?:^|\s)(#|@)[\w\-]+(?=\s*$)"

# State coordinate ranges
STATE_LON_RANGES = {
    "AL": (-88.473227, -84.889080), "AK": (-179.148909, 179.778470), "AZ": (-114.816510, -109.045223),
    "AR": (-94.617919, -89.644395), "CA": (-124.409591, -114.131211), "CO": (-109.060253, -102.041524),
    "CT": (-73.727775, -71.786994), "DE": (-75.788658, -75.048939), "FL": (-87.634896, -80.031056),
    "GA": (-85.605165, -80.840141), "HI": (-178.334698, -154.806773), "ID": (-117.243027, -111.043564),
    "IL": (-91.513079, -87.494756), "IN": (-88.097892, -84.787981), "IA": (-96.639704, -90.140061),
    "KS": (-102.051744, -94.588413), "KY": (-89.571510, -81.964971), "LA": (-94.043147, -88.817017),
    "ME": (-71.083924, -66.949895), "MD": (-79.487651, -75.048939), "MA": (-73.508142, -69.928393),
    "MI": (-90.418136, -82.413474), "MN": (-97.239209, -89.483385), "MS": (-91.655009, -88.097892),
    "MO": (-95.774704, -89.098843), "MT": (-116.050002, -104.039138), "NE": (-104.053514, -95.308290),
    "NV": (-120.005746, -114.039648), "NH": (-72.557247, -70.610621), "NJ": (-75.559614, -73.893979),
    "NM": (-109.050173, -103.001964), "NY": (-79.762152, -71.856214), "NC": (-84.321869, -75.460621),
    "ND": (-104.048900, -96.554507), "OH": (-84.820159, -80.518693), "OK": (-103.002455, -94.430662),
    "OR": (-124.566244, -116.463262), "PA": (-80.519891, -74.689516), "RI": (-71.886819, -71.120557),
    "SC": (-83.353910, -78.541138), "SD": (-104.057698, -96.436589), "TN": (-90.310298, -81.646900),
    "TX": (-106.645646, -93.508292), "UT": (-114.052998, -109.041058), "VT": (-73.437740, -71.464555),
    "VA": (-83.675395, -75.242266), "WA": (-124.763068, -116.915989), "WV": (-82.644739, -77.719519),
    "WI": (-92.889433, -86.763983), "WY": (-111.056888, -104.052160), "DC": (-77.119759, -76.909393),
    "PR": (-67.945404, -65.220703), "VI": (-65.013029, -64.564907), "GU": (144.618068, 144.956706),
    "AS": (-170.841600, -169.406622), "MP": (145.128345, 145.853700)
}
STATE_LAT_RANGES = {
    "AL": (30.137521, 35.008028), "AK": (51.214183, 71.538800), "AZ": (31.332177, 37.004260),
    "AR": (33.004106, 36.499749), "CA": (32.534156, 42.009518), "CO": (36.993076, 41.003444),
    "CT": (40.980144, 42.050587), "DE": (38.451013, 39.839007), "FL": (24.396308, 31.001056),
    "GA": (30.355657, 35.001180), "HI": (18.911680, 28.517269), "ID": (41.988057, 49.001146),
    "IL": (36.970298, 42.508481), "IN": (37.771743, 41.760592), "IA": (40.375437, 43.501196),
    "KS": (36.993016, 40.003166), "KY": (36.496486, 39.147458), "LA": (28.928609, 33.019543),
    "ME": (42.977764, 47.459686), "MD": (37.911717, 39.723043), "MA": (41.238100, 42.886589),
    "MI": (41.696118, 48.305884), "MN": (43.499361, 49.384358), "MS": (30.173943, 35.005002),
    "MO": (35.995683, 40.613639), "MT": (44.358209, 49.001390), "NE": (39.999932, 43.001708),
    "NV": (35.001857, 42.002207), "NH": (42.697037, 45.305476), "NJ": (38.928609, 41.357423),
    "NM": (31.332301, 37.000293), "NY": (40.496103, 45.015850), "NC": (33.840233, 36.588117),
    "ND": (45.935054, 49.000574), "OH": (38.403202, 41.977523), "OK": (33.615833, 37.002312),
    "OR": (41.991794, 46.299099), "PA": (39.719799, 42.269314), "RI": (41.146339, 42.018798),
    "SC": (32.034600, 35.215402), "SD": (42.479635, 45.945450), "TN": (34.982972, 36.678118),
    "TX": (25.837377, 36.500704), "UT": (36.997968, 42.001567), "VT": (42.726853, 45.016659),
    "VA": (36.540759, 39.466012), "WA": (45.543541, 49.002494), "WV": (37.201483, 40.638801),
    "WI": (42.491983, 47.080621), "WY": (40.994746, 45.005904), "DC": (38.791645, 38.995548),
    "PR": (17.926405, 18.516726), "VI": (17.673976, 18.412655), "GU": (13.234189, 13.654383),
    "AS": (-14.548699, -14.120151), "MP": (14.093068, 20.553762)
}

# Excel cell styles
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
PINK_FILL = PatternFill(start_color="FFC1CC", end_color="FFC1CC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Data type specifications
DTYPE_DICT = {
    "OrigRowNum": "int32",
    "customer": "string",
    "lat": "float64",
    "lon": "float64",
    "address": "string",
    "city": "string",
    "state": "string",
    "zip": "string",
    "download": "float64",
    "upload": "float64",
    "voip_lines_quantity": "int32",
    "business_customer": "int32",
    "technology": "string"
}

# Column name mapping for common variations
COLUMN_NAME_MAPPING = {
    # ZIP variations
    "zipcode": "zip",
    "zip code": "zip", 
    "zip_code": "zip",
    "postal code": "zip",
    "postal_code": "zip",
    "postalcode": "zip",
    # Longitude variations
    "longitude": "lon",
    "long": "lon",
    "lng": "lon",
    "lon_deg": "lon",
    "longitude_deg": "lon",
    "x": "lon",
    # Latitude variations
    "latitude": "lat",
    "lat_deg": "lat", 
    "latitude_deg": "lat",
    "y": "lat",
    # Customer variations
    "customer_id": "customer",
    "cust_id": "customer",
    # Address variations
    "street_address": "address",
    "addr": "address",
    "address1": "address",
    "address_1": "address",
    # State variations
    "region": "state",
    "st": "state",
    # Download variations
    "download_speed": "download",
    "down": "download",
    # Upload variations
    "up": "upload",
    "up_speed": "upload",
    # VoIP variations
    "voip": "voip_lines_quantity",
    "voip_lines": "voip_lines_quantity",
    "phone": "voip_lines_quantity",
    "phone_lines": "voip_lines_quantity",
    "lines": "voip_lines_quantity",
    # Technology variations
    "tech": "technology"
}

# File validation thresholds for determining valid vs invalid subscriber files
# Based on subscriber count and allowable percentage of address field errors
VALIDATION_THRESHOLDS = [
    {
        "min_subscribers": 0,
        "max_subscribers": 200,
        "max_error_percentage": 0.0,
        "description": "Small files: Zero tolerance for address errors"
    },
    {
        "min_subscribers": 201,
        "max_subscribers": 500,
        "max_error_percentage": 3.0,
        "description": "Medium files: Up to 3% address errors allowed"
    },
    {
        "min_subscribers": 501,
        "max_subscribers": 1500,
        "max_error_percentage": 2.0,
        "description": "Large files: Up to 2% address errors allowed"
    },
    {
        "min_subscribers": 1501,
        "max_subscribers": 999999,
        "max_error_percentage": 1.0,
        "description": "Very large files: Up to 1% address errors allowed"
    }
]

# Address field column names for validation assessment
# NOTE: state is NOT included because invalid state values (like "1087 BIA")
# usually indicate column misalignment or critical data errors, not just bad addresses
ADDRESS_COLUMNS = ["address", "city", "zip"]

def get_validation_threshold(subscriber_count):
    """
    Get the validation threshold configuration for a given subscriber count.
    
    Args:
        subscriber_count (int): Number of subscribers in the file
        
    Returns:
        dict: Threshold configuration with keys:
            - min_subscribers: Minimum count for this range
            - max_subscribers: Maximum count for this range  
            - max_error_percentage: Maximum allowable error percentage
            - description: Human-readable description
            
    Raises:
        ValueError: If subscriber_count is negative or no matching threshold found
    """
    if subscriber_count < 0:
        raise ValueError(f"Subscriber count cannot be negative: {subscriber_count}")
    
    for threshold in VALIDATION_THRESHOLDS:
        if threshold["min_subscribers"] <= subscriber_count <= threshold["max_subscribers"]:
            return threshold
    
    # Fallback - should not happen with current configuration
    raise ValueError(f"No validation threshold found for subscriber count: {subscriber_count}")


def is_address_column(column_name):
    """
    Check if a column name is considered an address field.
    
    Args:
        column_name (str): Name of the column to check
        
    Returns:
        bool: True if the column is an address field, False otherwise
    """
    return column_name.lower().strip() in [col.lower() for col in ADDRESS_COLUMNS]